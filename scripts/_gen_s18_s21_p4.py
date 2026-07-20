#!/usr/bin/env python3
"""Generate S18–S21 V3 PHASE 4 section TS files with verified Python outputs.

Authority: LANE-S18-S21-P4 Author.
Does not touch seed/checkpoint/ledger/orchestration.
Platform ids preserved: data-engineering, databases-orm, rag, fastapi.
V3 themes: EDA/incertidumbre | viz accesible | Excel factory | reportes trazables (CP-N2-B).
Español peruano; datos sintéticos only.
"""
from __future__ import annotations

import json
import subprocess
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

ROOT = Path(__file__).resolve().parents[1]
SECTIONS = ROOT / "src/lib/course/sections"
STATE = ROOT / "course-state"
LANES = STATE / "lanes"
LANE_ID = "LANE-S18-S21-P4"


def run_py(code: str) -> str:
    r = subprocess.run(
        [sys.executable, "-c", code],
        capture_output=True,
        text=True,
        timeout=90,
    )
    if r.returncode != 0:
        raise RuntimeError(f"Python failed:\nCODE:\n{code}\nSTDERR:\n{r.stderr}")
    out = r.stdout
    if out.endswith("\n"):
        out = out[:-1]
    return out


def esc(s: str) -> str:
    return s.replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")


def ts_str(s: str) -> str:
    return json.dumps(s, ensure_ascii=False)


@dataclass
class Ex:
    eid: str
    sub: str
    kind: str
    instruction: str
    hint1: str
    hint2: str
    starter: str
    solution: str
    edge: list[str] = field(default_factory=list)
    tests: str = ""
    feedback: str = ""
    title: str = "exercise.py"
    output: str = ""


@dataclass
class Demo:
    did: str
    sub: str
    description: str
    code: str
    why: str
    title: str = "demo.py"
    output: str = ""
    env: str = "local-python"


@dataclass
class Theory:
    heading: str
    paragraphs: list[str]
    sub: Optional[str] = None
    code: Optional[str] = None
    code_title: str = "example.py"
    code_out: str = ""
    callout: Optional[tuple[str, str, str]] = None


def verify_all(demos: list[Demo], exercises: list[Ex]) -> dict[str, str]:
    log: dict[str, str] = {}
    for d in demos:
        d.output = run_py(d.code)
        log[d.did] = "VERIFIED"
    for e in exercises:
        e.output = run_py(e.solution)
        log[e.eid] = "VERIFIED"
    return log


def theory_block(t: Theory) -> str:
    paras = ",\n".join(f"        {ts_str(p)}" for p in t.paragraphs)
    lines = [
        "    {",
        f"      heading: {ts_str(t.heading)},",
    ]
    if t.sub:
        lines.append(f"      subtopicId: {ts_str(t.sub)},")
    lines.append(f"      paragraphs: [\n{paras},\n      ],")
    if t.code is not None:
        out = t.code_out if t.code_out else run_py(t.code)
        t.code_out = out
        lines.append("      code: {")
        lines.append("        language: 'python',")
        lines.append(f"        title: {ts_str(t.code_title)},")
        lines.append(f"        code: `{esc(t.code)}`,")
        lines.append(f"        output: `{esc(out)}`,")
        lines.append("      },")
    if t.callout:
        ctype, ctitle, ccontent = t.callout
        lines.append("      callout: {")
        lines.append(f"        type: {ts_str(ctype)},")
        lines.append(f"        title: {ts_str(ctitle)},")
        lines.append(f"        content:\n          {ts_str(ccontent)},")
        lines.append("      },")
    lines.append("    },")
    return "\n".join(lines)


def demo_step(d: Demo) -> str:
    if not d.output:
        d.output = run_py(d.code)
    return f"""      {{
        demoId: {ts_str(d.did)},
        subtopicId: {ts_str(d.sub)},
        environment: {ts_str(d.env)},
        description: {ts_str(d.description)},
        code: {{
          language: 'python',
          title: {ts_str(d.title)},
          code: `{esc(d.code)}`,
          output: `{esc(d.output)}`,
        }},
        why: {ts_str(d.why)},
      }},"""


def exercise_step(e: Ex) -> str:
    if not e.output:
        e.output = run_py(e.solution)
    edges = ", ".join(ts_str(x) for x in e.edge) if e.edge else ""
    return f"""      {{
        id: {ts_str(e.eid)},
        subtopicId: {ts_str(e.sub)},
        kind: {ts_str(e.kind)},
        instruction:
          {ts_str(e.instruction)},
        hint: {ts_str(e.hint1)},
        hints: [
          {ts_str(e.hint1)},
          {ts_str(e.hint2)},
        ],
        edgeCases: [{edges}],
        tests: {ts_str(e.tests or "salida coincide con solution output")},
        feedback: {ts_str(e.feedback or "Compara tu salida con la solución.")},
        starterCode: {{
          language: 'python',
          title: {ts_str(e.title)},
          code: `{esc(e.starter)}`,
        }},
        solutionCode: {{
          language: 'python',
          title: {ts_str(e.title)},
          code: `{esc(e.solution)}`,
          output: `{esc(e.output)}`,
        }},
      }},"""


def build_section(
    export_name: str,
    meta: dict[str, Any],
    theories: list[Theory],
    demos: list[Demo],
    exercises: list[Ex],
    i_do_intro: str,
    we_do_intro: str,
    youdo: dict[str, Any],
    selfcheck: list[dict],
    resources: dict,
) -> tuple[str, dict[str, str]]:
    for t in theories:
        if t.code:
            t.code_out = run_py(t.code)
    log = verify_all(demos, exercises)

    parts: list[str] = []
    parts.append("import type { CourseSection } from '../../types'")
    parts.append("")
    parts.append(f"export const {export_name}: CourseSection = {{")
    for k in [
        "id",
        "index",
        "title",
        "shortTitle",
        "tagline",
        "estimatedHours",
        "level",
        "phase",
        "icon",
        "accentColor",
    ]:
        v = meta[k]
        if isinstance(v, str):
            parts.append(f"  {k}: {ts_str(v)},")
        else:
            parts.append(f"  {k}: {v},")
    parts.append(f"  jobRelevance:\n    {ts_str(meta['jobRelevance'])},")
    parts.append("  learningOutcomes: [")
    for lo in meta["learningOutcomes"]:
        parts.append(f"    {{ text: {ts_str(lo)} }},")
    parts.append("  ],")
    parts.append("  theory: [")
    for t in theories:
        parts.append(theory_block(t))
    parts.append("  ],")
    parts.append("  iDo: {")
    parts.append(f"    intro: {ts_str(i_do_intro)},")
    parts.append("    steps: [")
    for d in demos:
        parts.append(demo_step(d))
    parts.append("    ],")
    parts.append("  },")
    parts.append("  weDo: {")
    parts.append(f"    intro: {ts_str(we_do_intro)},")
    parts.append("    steps: [")
    for e in exercises:
        parts.append(exercise_step(e))
    parts.append("    ],")
    parts.append("  },")
    parts.append("  youDo: {")
    parts.append(f"    title: {ts_str(youdo['title'])},")
    parts.append(f"    context:\n      {ts_str(youdo['context'])},")
    parts.append("    objectives: [")
    for o in youdo["objectives"]:
        parts.append(f"      {ts_str(o)},")
    parts.append("    ],")
    parts.append("    requirements: [")
    for r in youdo["requirements"]:
        parts.append(f"      {ts_str(r)},")
    parts.append("    ],")
    parts.append(f"    starterCode: `{esc(youdo['starterCode'])}`,")
    parts.append(f"    portfolioNote:\n      {ts_str(youdo['portfolioNote'])},")
    parts.append("    rubric: [")
    for item in youdo["rubric"]:
        parts.append(
            f"      {{ criterion: {ts_str(item['criterion'])}, weight: {ts_str(item['weight'])} }},"
        )
    parts.append("    ],")
    parts.append("  },")
    parts.append("  selfCheck: {")
    parts.append("    questions: [")
    for q in selfcheck:
        opts = ",\n".join(f"          {ts_str(o)}" for o in q["options"])
        parts.append("      {")
        parts.append(f"        question: {ts_str(q['question'])},")
        parts.append(f"        options: [\n{opts},\n        ],")
        parts.append(f"        correctIndex: {q['correctIndex']},")
        parts.append(f"        explanation:\n          {ts_str(q['explanation'])},")
        parts.append("      },")
    parts.append("    ],")
    parts.append("  },")
    parts.append("  resources: {")
    parts.append("    docs: [")
    for d in resources["docs"]:
        note = f",\n        note: {ts_str(d['note'])}" if d.get("note") else ""
        parts.append(
            f"      {{\n        label: {ts_str(d['label'])},\n        url: {ts_str(d['url'])}{note},\n      }},"
        )
    parts.append("    ],")
    parts.append("    books: [")
    for b in resources["books"]:
        parts.append(
            f"      {{\n        label: {ts_str(b['label'])},\n        note: {ts_str(b['note'])},\n      }},"
        )
    parts.append("    ],")
    parts.append("    courses: [")
    for c in resources["courses"]:
        note = f",\n        note: {ts_str(c['note'])}" if c.get("note") else ""
        parts.append(
            f"      {{\n        label: {ts_str(c['label'])},\n        url: {ts_str(c['url'])}{note},\n      }},"
        )
    parts.append("    ],")
    parts.append("  },")
    parts.append("}")
    parts.append("")
    return "\n".join(parts), log


def progress_payload(
    section: str,
    section_id: str,
    file_ts: str,
    meta: dict,
    slugs: list[tuple[str, str]],
    log: dict[str, str],
    youdo_title: str,
) -> dict:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    subtopics = {}
    for sid, slug in slugs:
        subtopics[sid] = {
            "slug": slug,
            "status": "COMPLETE",
            "theory": True,
            "demo": f"{sid}-DEMO",
            "exercises": [f"{sid}-E1", f"{sid}-E2", f"{sid}-E3"],
            "executed": True,
        }
    items = [
        {"id": k, "status": v, "notes": "python3 executed; output embedded"}
        for k, v in log.items()
    ]
    return {
        "version": "3.2",
        "section": section,
        "section_id": section_id,
        "phase": "PHASE_4_COMPLETE",
        "lane": LANE_ID,
        "generated_at": now,
        "prompt_version": "3.2",
        "authority": (
            "PARALLEL_PRODUCTION author agent — does not mark section passed; "
            "does not edit checkpoint/ledger/seed/orchestration"
        ),
        "preamble": {
            "objective": f"Retarget {section_id} section TS to V3 roadmap; author full packages for all 8 subtopics.",
            "scope_in": [
                f"src/lib/course/sections/{file_ts}",
                f"course-state/{section.lower()}_phase4_progress.json",
                f"course-state/lanes/{LANE_ID}.status.json",
            ],
            "scope_out": [
                "checkpoint.json",
                "section_ledger.json",
                "issue_registry.json",
                "parallel_orchestration.json",
                "prisma/seed.ts",
                "PHASE 5 exam A/B/C stems",
                "section_passed",
            ],
        },
        "meta_updates": meta,
        "subtopics_authored": subtopics,
        "counts": {
            "subtopics_done": 8,
            "subtopics_target": 8,
            "demos_done": 8,
            "demos_target": 8,
            "exercises_done": 24,
            "exercises_target": 24,
            "hints_per_exercise": 2,
            "theory_blocks_with_subtopicId": 8,
            "intro_theory_blocks": 1,
        },
        "youDo": youdo_title,
        "execution_log": {
            "runtime": "python3 local",
            "executed_at": now[:10],
            "UNVERIFIED": [],
            "items": items,
        },
        "section_passed": False,
        "locale": "es-PE",
        "synthetic_data_only": True,
    }


RUBRIC_STD = [
    {"criterion": "Alineación al gate V3 de la sección", "weight": "25%"},
    {"criterion": "Correctitud técnica en entorno declarado", "weight": "20%"},
    {"criterion": "Privacidad / sin PII real / sin secretos", "weight": "20%"},
    {"criterion": "Pruebas o casos de borde documentados", "weight": "15%"},
    {"criterion": "Código legible y límites claros", "weight": "10%"},
    {"criterion": "Documentación en español profesional", "weight": "10%"},
]


def ex(eid, sub, kind, instr, h1, h2, starter, solution, edge):
    return Ex(eid, sub, kind, instr, h1, h2, starter, solution, edge=edge)


# ---------------------------------------------------------------------------
# S18 — EDA, estadística descriptiva e incertidumbre (platform id: data-engineering)
# ---------------------------------------------------------------------------


def build_s18() -> tuple[str, dict[str, str], list[tuple[str, str]], dict, str]:
    slugs = [
        ("S18-T1-A", "center-spread-quantiles"),
        ("S18-T1-B", "robust-metrics-scales"),
        ("S18-T2-A", "population-sample-bias"),
        ("S18-T2-B", "intervals-effect-size"),
        ("S18-T3-A", "correlation-confounds"),
        ("S18-T3-B", "segments-anomalies-no-causal"),
        ("S18-T4-A", "questions-hypotheses-evidence"),
        ("S18-T4-B", "repro-notebook-data-notes"),
    ]
    meta = {
        "id": "data-engineering",
        "index": 18,
        "title": "EDA, estadística descriptiva e incertidumbre",
        "shortTitle": "EDA e incertidumbre",
        "tagline": "EDA que diferencia hallazgo, hipótesis y decisión; cada conclusión referencia un cálculo y declara incertidumbre",
        "estimatedHours": 12,
        "level": "Competente",
        "phase": 1,
        "icon": "Wrench",
        "accentColor": "bg-gradient-to-br from-blue-500 to-indigo-600",
        "jobRelevance": (
            "En analytics y data products de banca, fintech y retail en Perú, un **EDA honesto** separa hallazgo, "
            "hipótesis y decisión. Esta sección (id de plataforma `data-engineering` conservado) retematiza a V3 "
            "**EDA, estadística descriptiva e incertidumbre** e inicia **CP-N2-B** con datos sintéticos, intervalos "
            "básicos y data notes reproducibles."
        ),
        "learningOutcomes": [
            "Resumir distribuciones con centro, dispersión y cuantiles",
            "Elegir métricas robustas y escalas honestas",
            "Identificar sesgo de población/muestra",
            "Reportar intervalos y tamaños de efecto básicos",
            "Interpretar correlación sin confundir causalidad",
            "Segmentar y marcar anomalías con límites claros",
            "Estructurar preguntas, hipótesis y evidencia",
            "Entregar notebook/script reproducible con data notes",
        ],
    }

    theories = [
        Theory(
            heading="De “Ingeniería de Datos Intermedia” a EDA e incertidumbre (mapa de la sección)",
            paragraphs=[
                "En V3, **S18 no es el path principal de Prefect, Parquet ni Great Expectations**. Ese material se reubica conceptualmente hacia ingeniería de datos avanzada. Aquí construyes el **inicio de CP-N2-B**: centro/dispersión, métricas robustas, sesgo muestral, intervalos, correlación sin causalidad y notebooks con data notes.",
                "El hilo conductor es un **dataset sintético de tickets/montos** (regiones Lima/Arequipa/Cusco, ids `T00x`). Cada hallazgo cita un cálculo y declara incertidumbre.",
                "Orden: **T1 Distribuciones** → **T2 Inferencia básica** → **T3 Relaciones** → **T4 Comunicación de evidencia**.",
            ],
            callout=(
                "info",
                "Contenido reubicado conceptualmente",
                "Material legado de Prefect/Parquet/GE de este archivo **no es el camino V3 del estudiante en S18**. Target: EDA e incertidumbre para CP-N2-B (inicio). Solo datos sintéticos; nunca PII real.",
            ),
        ),
        Theory(
            heading="centro, dispersión y cuantiles",
            sub="S18-T1-A",
            paragraphs=[
                "El **centro** se resume con media (`mean`) o mediana (`median`). La **dispersión** con desviación estándar (`std`) o rango intercuartílico (**IQR** = Q3−Q1).",
                "Los **cuantiles** (p25, p50, p75, p90) describen la forma sin asumir normalidad. En montos de tickets, la media se mueve con outliers; la mediana suele ser más estable.",
                "Reporta siempre n, y preferiblemente min/max o un cuantil alto para contexto de cola.",
            ],
            code="""import numpy as np

montos = np.array([12.5, 18.0, 22.0, 25.5, 30.0, 45.0, 120.0])  # sintético PEN
print("n", montos.size)
print("mean", round(float(montos.mean()), 2))
print("median", float(np.median(montos)))
print("std", round(float(montos.std(ddof=1)), 2))
q = np.quantile(montos, [0.25, 0.5, 0.75, 0.9])
print("q25_q50_q75_q90", [round(float(x), 2) for x in q])
print("IQR", round(float(q[2] - q[0]), 2))""",
            code_title="center_spread.py",
            callout=(
                "tip",
                "ddof en std muestral",
                "Para muestra usa ddof=1 (n−1). Documenta si usas población (ddof=0).",
            ),
        ),
        Theory(
            heading="métricas robustas y escalas",
            sub="S18-T1-B",
            paragraphs=[
                "Métricas **robustas**: mediana, IQR, MAD (median absolute deviation). Resisten outliers mejor que media/std.",
                "La **escala** importa: log1p de montos reduce asimetría visual; no inventes “crecimiento %” sobre ejes engañosos.",
                "Elige métrica según la pregunta: “ticket típico” → mediana; “ingreso total esperado” → media (con cola documentada).",
            ],
            code="""import numpy as np

x = np.array([10.0, 12.0, 11.0, 13.0, 12.5, 200.0])
med = float(np.median(x))
mad = float(np.median(np.abs(x - med)))
print("median", med, "MAD", mad)
print("mean_vs_median", round(float(x.mean()), 2), med)
print("log1p", np.round(np.log1p(x), 3).tolist())""",
            code_title="robust_scale.py",
            callout=(
                "warning",
                "Escala log y comunicación",
                "Si usas log, dilo en el eje y en la conclusión; no compares diferencias log como soles PEN sin transformar.",
            ),
        ),
        Theory(
            heading="población, muestra y sesgo",
            sub="S18-T2-A",
            paragraphs=[
                "La **población** es el universo de interés; la **muestra** es lo observado. El **sesgo de selección** aparece si el muestreo no es representativo (p. ej. solo Lima).",
                "Compara distribución de la muestra vs marco conocido (cuotas por región). Documenta exclusiones (filtros de fecha, canal).",
                "Sin marco poblacional, declara **cobertura limitada** y no generalices a “todos los clientes del Perú”.",
            ],
            code="""import numpy as np

# población sintética de tickets por región
pob = {"Lima": 0.55, "Arequipa": 0.25, "Cusco": 0.20}
# muestra sesgada: sobremuestra Lima
muestra = np.array(["Lima"] * 40 + ["Arequipa"] * 8 + ["Cusco"] * 2)
from collections import Counter
c = Counter(muestra)
n = len(muestra)
share = {k: round(v / n, 3) for k, v in c.items()}
print("share_muestra", share)
print("share_pob", pob)
bias = {k: round(share.get(k, 0) - pob[k], 3) for k in pob}
print("bias_pp", bias)""",
            code_title="sample_bias.py",
            callout=(
                "danger",
                "Sesgo ≠ error de cálculo",
                "Un mean correcto sobre una muestra sesgada sigue siendo una estimación sesgada de la población.",
            ),
        ),
        Theory(
            heading="intervalos básicos y tamaño de efecto",
            sub="S18-T2-B",
            paragraphs=[
                "Un **intervalo de confianza** (IC) aproximado para la media con n grande: media ± z * (s/√n). Para n pequeño o no normal, sé cauteloso y reporta bootstrap simple si aplica.",
                "El **tamaño de efecto** (p. ej. diferencia de medias / s pooled, o diferencia de medianas) comunica magnitud, no solo “significativo”.",
                "Nunca digas “probado” con un solo IC; di “compatible con” y reporta n.",
            ],
            code="""import numpy as np

rng = np.random.default_rng(7)
a = rng.normal(100, 15, size=40)
b = rng.normal(108, 15, size=40)
ma, mb = a.mean(), b.mean()
sa, sb = a.std(ddof=1), b.std(ddof=1)
# IC 95% approx para media de b
se = sb / np.sqrt(len(b))
ic = (mb - 1.96 * se, mb + 1.96 * se)
sp = np.sqrt((sa**2 + sb**2) / 2)
d = (mb - ma) / sp
print("mean_a", round(ma, 2), "mean_b", round(mb, 2))
print("ic95_b", (round(ic[0], 2), round(ic[1], 2)))
print("cohens_d", round(float(d), 3))""",
            code_title="interval_effect.py",
            callout=(
                "tip",
                "Efecto + incertidumbre",
                "Reporta diferencia puntual, IC y n. El tamaño de efecto evita obsesionarse solo con p-values.",
            ),
        ),
        Theory(
            heading="correlación y confusión",
            sub="S18-T3-A",
            paragraphs=[
                "La **correlación** (Pearson/Spearman) mide asociación lineal/monótona, no causa. Un confounder puede crear asociación espuria.",
                "Antes de “X causa Y”, lista causas comunes y diseños que las romperían (experimentos, instrumentos). En EDA, etiqueta como **asociación observada**.",
                "Pearson es sensible a outliers; Spearman usa rangos y es más robusto a monótonas no lineales leves.",
            ],
            code="""import numpy as np

rng = np.random.default_rng(1)
# confounder Z genera X e Y
z = rng.normal(0, 1, 80)
x = 0.8 * z + rng.normal(0, 0.3, 80)
y = 0.7 * z + rng.normal(0, 0.3, 80)
r = np.corrcoef(x, y)[0, 1]
print("pearson_xy", round(float(r), 3))
# control parcial tosco: residualizar Z
def resid(a, z):
    b = np.polyfit(z, a, 1)
    return a - (b[0] * z + b[1])
rx, ry = resid(x, z), resid(y, z)
print("pearson_residual", round(float(np.corrcoef(rx, ry)[0, 1]), 3))""",
            code_title="corr_confound.py",
            callout=(
                "warning",
                "Correlación ≠ causalidad",
                "Si no controlas confusores ni tienes diseño causal, no uses verbos causales en el informe.",
            ),
        ),
        Theory(
            heading="segmentación, anomalías y causalidad no demostrada",
            sub="S18-T3-B",
            paragraphs=[
                "Segmenta por región, canal o cohorte con **reglas explícitas**. Las anomalías (p. ej. Tukey: fuera de [Q1−1.5·IQR, Q3+1.5·IQR]) son candidatos a revisión, no “fraudes demostrados”.",
                "Marca flags y tasas por segmento; evita claims causales del tipo “Cusco genera outliers porque…”.",
                "Documenta umbral, n por segmento y si el método es univariado.",
            ],
            code="""import numpy as np
import pandas as pd

df = pd.DataFrame({
    "region": ["Lima"] * 5 + ["Cusco"] * 5,
    "monto": [20, 22, 21, 25, 19, 18, 23, 80, 21, 20],
})
q1, q3 = df["monto"].quantile(0.25), df["monto"].quantile(0.75)
iqr = q3 - q1
lo, hi = q1 - 1.5 * iqr, q3 + 1.5 * iqr
df["anomalia"] = (df["monto"] < lo) | (df["monto"] > hi)
print("limites", round(float(lo), 2), round(float(hi), 2))
print(df.groupby("region")["anomalia"].mean().round(3).to_dict())
print("ids_anom", df.index[df["anomalia"]].tolist())""",
            code_title="segments_anom.py",
            callout=(
                "info",
                "Sin claim causal",
                "Una tasa mayor de anomalías en un segmento es hallazgo descriptivo, no prueba de causa.",
            ),
        ),
        Theory(
            heading="preguntas, hipótesis y evidencia",
            sub="S18-T4-A",
            paragraphs=[
                "Separa tres capas: **pregunta de negocio**, **hipótesis comprobable**, **evidencia calculada**. El hallazgo no es la decisión.",
                "Plantilla: Pregunta → Métrica → Resultado (n, punto, IC) → Límite → Siguiente paso.",
                "En CP-N2-B, cada slide o celda de insights debe poder rastrearse a un cálculo en el script.",
            ],
            code="""# plantilla de traza hallazgo → cálculo (sintético)
pregunta = "¿El ticket mediano en Lima supera 25 PEN?"
metrica = "median(monto | region==Lima)"
resultado = {"n": 40, "median": 27.5, "ic_boot_approx": (24.0, 31.0)}
limite = "muestra de canal web, no incluye tienda física"
decision_sugerida = "hipótesis provisional: sí en web Lima; validar con marco completo"
print("pregunta:", pregunta)
print("metrica:", metrica)
print("resultado:", resultado)
print("limite:", limite)
print("no_es_decision:", "no lanzar campaña aún")""",
            code_title="qhe_template.py",
            callout=(
                "tip",
                "Hallazgo ≠ decisión",
                "La decisión de negocio requiere costo, riesgo y cobertura; el EDA solo aporta evidencia parcial.",
            ),
        ),
        Theory(
            heading="notebook reproducible y data notes",
            sub="S18-T4-B",
            paragraphs=[
                "Un **data note** documenta origen, fecha de corte, filtros, n final y exclusiones. El script debe reejecutarse con seed fijo.",
                "Fija versiones (pandas/numpy), rutas relativas y outputs en carpeta `out/`. Evita celdas que muten estado sin reorden claro.",
                "Checklist: seed, schema, n pre/post filtros, hash o conteo de filas, límites de generalización.",
            ],
            code="""import hashlib
import json
import numpy as np
import pandas as pd

df = pd.DataFrame({"ticket_id": ["T001", "T002", "T003"], "monto": [10.0, 20.0, 15.0]})
payload = df.to_csv(index=False).encode()
note = {
    "origen": "sintetico_local",
    "corte": "2024-06-30",
    "n": len(df),
    "filtros": ["monto > 0"],
    "seed": 42,
    "row_sha1_8": hashlib.sha1(payload).hexdigest()[:8],
}
print(json.dumps(note, ensure_ascii=False))
print("mean", float(df["monto"].mean()))""",
            code_title="data_notes.py",
            callout=(
                "success",
                "Reproducibilidad mínima",
                "Si otro agente no puede regenerar los mismos n y métricas clave, el notebook no cierra el gate S18.",
            ),
        ),
    ]

    demos = [
        Demo(
            "S18-T1-A-DEMO",
            "S18-T1-A",
            "Resumir distribución de montos sintéticos con centro, dispersión y cuantiles",
            """import numpy as np

rng = np.random.default_rng(18)
montos = np.concatenate([rng.lognormal(3.0, 0.4, 90), np.array([400.0, 450.0])])
def resumen(x):
    q = np.quantile(x, [0.25, 0.5, 0.75, 0.9])
    return {
        "n": int(x.size),
        "mean": round(float(x.mean()), 2),
        "median": round(float(np.median(x)), 2),
        "std": round(float(x.std(ddof=1)), 2),
        "q25": round(float(q[0]), 2),
        "q50": round(float(q[1]), 2),
        "q75": round(float(q[2]), 2),
        "q90": round(float(q[3]), 2),
        "IQR": round(float(q[2] - q[0]), 2),
    }
print(resumen(montos))""",
            "Un resumen tabular con n y cuantiles es la base de cualquier hallazgo de distribución.",
            title="demo_center_spread.py",
        ),
        Demo(
            "S18-T1-B-DEMO",
            "S18-T1-B",
            "Comparar media vs mediana/MAD y escala log1p en montos con outlier",
            """import numpy as np

x = np.array([15, 16, 14, 18, 17, 16, 15, 200], dtype=float)
med = float(np.median(x))
mad = float(np.median(np.abs(x - med)))
print("mean", round(float(x.mean()), 2))
print("median", med, "MAD", mad)
print("ratio_mean_median", round(float(x.mean()) / med, 2))
print("log1p_median", round(float(np.median(np.log1p(x))), 3))""",
            "Cuando mean ≫ median, prioriza métricas robustas y declara la cola.",
            title="demo_robust.py",
        ),
        Demo(
            "S18-T2-A-DEMO",
            "S18-T2-A",
            "Diagnosticar sesgo de muestreo por región frente a cuotas poblacionales",
            """from collections import Counter
import numpy as np

pob = {"Lima": 0.50, "Arequipa": 0.30, "Cusco": 0.20}
muestra = ["Lima"] * 70 + ["Arequipa"] * 20 + ["Cusco"] * 10
c = Counter(muestra)
n = sum(c.values())
share = {k: c[k] / n for k in pob}
print({k: round(share[k], 3) for k in pob})
print("max_abs_bias_pp", round(max(abs(share[k] - pob[k]) for k in pob), 3))
print("cobertura", "LIMITADA" if max(abs(share[k] - pob[k]) for k in pob) > 0.1 else "OK")""",
            "El sesgo de cuota se mide en puntos porcentuales, no solo con “se ve bien”.",
            title="demo_bias.py",
        ),
        Demo(
            "S18-T2-B-DEMO",
            "S18-T2-B",
            "Reportar IC 95% aproximado y Cohen's d entre dos grupos sintéticos",
            """import numpy as np

rng = np.random.default_rng(21)
ctrl = rng.normal(50, 10, 35)
trat = rng.normal(55, 10, 35)
diff = trat.mean() - ctrl.mean()
se = np.sqrt(ctrl.var(ddof=1)/len(ctrl) + trat.var(ddof=1)/len(trat))
ic = (diff - 1.96*se, diff + 1.96*se)
sp = np.sqrt((ctrl.var(ddof=1) + trat.var(ddof=1)) / 2)
d = diff / sp
print("diff", round(float(diff), 2))
print("ic95", (round(float(ic[0]), 2), round(float(ic[1]), 2)))
print("cohens_d", round(float(d), 3))
print("n", len(ctrl), len(trat))""",
            "Magnitud + intervalo + n comunican incertidumbre mejor que un solo p-value.",
            title="demo_effect.py",
        ),
        Demo(
            "S18-T3-A-DEMO",
            "S18-T3-A",
            "Mostrar correlación alta generada por confusor y caída al residualizar",
            """import numpy as np

rng = np.random.default_rng(3)
z = rng.normal(size=100)
x = z + rng.normal(scale=0.2, size=100)
y = z + rng.normal(scale=0.2, size=100)
r_raw = float(np.corrcoef(x, y)[0, 1])
bx = np.polyfit(z, x, 1)
by = np.polyfit(z, y, 1)
rx = x - (bx[0]*z + bx[1])
ry = y - (by[0]*z + by[1])
r_res = float(np.corrcoef(rx, ry)[0, 1])
print("r_raw", round(r_raw, 3))
print("r_residual_z", round(r_res, 3))
print("claim", "asociacion_observada_no_causal")""",
            "Enseña a no inferir causalidad solo porque r es alto.",
            title="demo_corr.py",
        ),
        Demo(
            "S18-T3-B-DEMO",
            "S18-T3-B",
            "Segmentar por región y marcar anomalías Tukey sin claim causal",
            """import pandas as pd

df = pd.DataFrame({
    "region": ["Lima"]*8 + ["Arequipa"]*6 + ["Cusco"]*6,
    "monto": [20,22,21,19,25,24,23,22, 18,19,20,21,17,55, 16,18,19,20,17,90],
})
q1, q3 = df["monto"].quantile([0.25, 0.75])
iqr = q3 - q1
lo, hi = q1 - 1.5*iqr, q3 + 1.5*iqr
df["flag"] = (df["monto"] < lo) | (df["monto"] > hi)
print("lo_hi", round(float(lo),2), round(float(hi),2))
print(df.groupby("region")["flag"].agg(["sum","mean"]).round(3).to_dict())
print("sin_claim_causal", True)""",
            "Flags + tasas por segmento; la narrativa causal queda fuera del EDA.",
            title="demo_segments.py",
        ),
        Demo(
            "S18-T4-A-DEMO",
            "S18-T4-A",
            "Separar pregunta, hipótesis, evidencia y no-decisión en un dict trazable",
            """evidencia = {
    "pregunta": "¿Hay diferencia de ticket mediano Lima vs Cusco?",
    "hipotesis": "mediana_Lima > mediana_Cusco en canal web junio",
    "calculo": "median por region, n>=30",
    "resultado": {"Lima": 28.0, "Cusco": 22.5, "n_Lima": 40, "n_Cusco": 32},
    "incertidumbre": "sin IC bootstrap en este corte; muestra web-only",
    "decision": None,
}
print(evidencia["pregunta"])
print("hallazgo", evidencia["resultado"])
print("decision_es_none", evidencia["decision"] is None)""",
            "La traza pregunta→cálculo→límite es el artefacto de calidad de CP-N2-B inicio.",
            title="demo_qhe.py",
        ),
        Demo(
            "S18-T4-B-DEMO",
            "S18-T4-B",
            "Generar data note con n, filtros, seed y huella de filas",
            """import hashlib, json
import pandas as pd

df = pd.DataFrame({
    "ticket_id": [f"T{i:03d}" for i in range(1, 6)],
    "monto": [10.0, 12.0, 0.0, 15.0, 11.0],
    "region": ["Lima", "Cusco", "Lima", "Arequipa", "Lima"],
})
n0 = len(df)
df2 = df[df["monto"] > 0].copy()
blob = df2.sort_values("ticket_id").to_csv(index=False).encode()
note = {
    "origen": "sintetico",
    "n_raw": n0,
    "n_final": len(df2),
    "filtros": ["monto > 0"],
    "seed": 18,
    "sha1_8": hashlib.sha1(blob).hexdigest()[:8],
}
print(json.dumps(note, ensure_ascii=False))
print("median_final", float(df2["monto"].median()))""",
            "Data notes hacen auditable el notebook del portfolio.",
            title="demo_datanote.py",
        ),
    ]

    exercises = [
        ex("S18-T1-A-E1", "S18-T1-A", "guided",
           "Con montos = [10, 12, 14, 16, 100], imprime n, mean redondeada a 2 decimales y median.",
           "Usa len o .size; np.mean y np.median.",
           "round(float(...), 2) para la media.",
           "import numpy as np\nmontos = np.array([10, 12, 14, 16, 100], dtype=float)\n# TODO\n",
           """import numpy as np
montos = np.array([10, 12, 14, 16, 100], dtype=float)
print("n", montos.size)
print("mean", round(float(montos.mean()), 2))
print("median", float(np.median(montos)))""",
           ["lista vacía", "todos iguales"]),
        ex("S18-T1-A-E2", "S18-T1-A", "independent",
           "Calcula Q1, Q3 e IQR de montos = [5, 8, 9, 10, 12, 13, 40] e imprímelos redondeados a 2 decimales.",
           "np.quantile con [0.25, 0.75].",
           "IQR = Q3 - Q1.",
           "import numpy as np\n# TODO\n",
           """import numpy as np
montos = np.array([5, 8, 9, 10, 12, 13, 40], dtype=float)
q1, q3 = np.quantile(montos, [0.25, 0.75])
print("Q1", round(float(q1), 2))
print("Q3", round(float(q3), 2))
print("IQR", round(float(q3 - q1), 2))""",
           ["n=1", "empates en cuantiles"]),
        ex("S18-T1-A-E3", "S18-T1-A", "transfer",
           "Escribe resumen(x) que devuelva dict con n, mean, median, std (ddof=1) redondeados; pruébalo con [1,2,3,4,5].",
           "std con ddof=1.",
           "Redondea mean/median/std a 4 decimales o 2 según print claro.",
           "import numpy as np\n# TODO resumen\n",
           """import numpy as np

def resumen(x):
    x = np.asarray(x, dtype=float)
    return {
        "n": int(x.size),
        "mean": round(float(x.mean()), 4),
        "median": round(float(np.median(x)), 4),
        "std": round(float(x.std(ddof=1)), 4),
    }
print(resumen([1, 2, 3, 4, 5]))""",
           ["array vacío debe fallar o manejarse"]),
        ex("S18-T1-B-E1", "S18-T1-B", "guided",
           "Para x=[10,11,12,13,100], imprime mean, median y la razón mean/median redondeada a 2.",
           "np.mean y np.median.",
           "ratio = mean/median.",
           "import numpy as np\n# TODO\n",
           """import numpy as np
x = np.array([10, 11, 12, 13, 100], dtype=float)
m = float(x.mean())
med = float(np.median(x))
print("mean", round(m, 2))
print("median", med)
print("ratio", round(m / med, 2))""",
           ["median 0"]),
        ex("S18-T1-B-E2", "S18-T1-B", "independent",
           "Calcula MAD de x=[2,3,4,5,100] respecto a la mediana e imprímelo.",
           "MAD = median(|x - median(x)|).",
           "Usa np.median y np.abs.",
           "import numpy as np\n# TODO\n",
           """import numpy as np
x = np.array([2, 3, 4, 5, 100], dtype=float)
med = float(np.median(x))
mad = float(np.median(np.abs(x - med)))
print("MAD", mad)""",
           ["todos iguales → MAD 0"]),
        ex("S18-T1-B-E3", "S18-T1-B", "transfer",
           "Imprime log1p de [0, 1, 9, 99] redondeado a 3 decimales como lista.",
           "np.log1p.",
           "tolist + round por elemento.",
           "import numpy as np\n# TODO\n",
           """import numpy as np
x = np.array([0, 1, 9, 99], dtype=float)
print([round(float(v), 3) for v in np.log1p(x)])""",
           ["negativos no válidos en log1p de montos"]),
        ex("S18-T2-A-E1", "S18-T2-A", "guided",
           "Dada muestra=['Lima','Lima','Cusco','Lima'] y pob Lima=0.5 Cusco=0.5, imprime share de Lima en la muestra.",
           "Cuenta ocurrencias / n.",
           "round a 2 decimales.",
           "# TODO\n",
           """muestra = ["Lima", "Lima", "Cusco", "Lima"]
share_lima = muestra.count("Lima") / len(muestra)
print("share_Lima", round(share_lima, 2))""",
           ["muestra vacía"]),
        ex("S18-T2-A-E2", "S18-T2-A", "independent",
           "Calcula bias_pp = share_muestra - share_pob para Lima si muestra tiene 8 Lima y 2 Cusco; pob Lima=0.5.",
           "share = 8/10.",
           "bias = share - 0.5.",
           "# TODO\n",
           """share = 8 / 10
pob = 0.5
print("bias_Lima_pp", round(share - pob, 2))""",
           ["regiones faltantes en muestra"]),
        ex("S18-T2-A-E3", "S18-T2-A", "transfer",
           "Escribe max_bias(pob, counts) que devuelva el máximo |share-pob| y pruébalo con pob Lima:0.5 Cusco:0.5 y counts Lima:9 Cusco:1.",
           "share = count/sum(counts).",
           "max de valores absolutos.",
           "# TODO\n",
           """def max_bias(pob, counts):
    n = sum(counts.values())
    return max(abs(counts[k] / n - pob[k]) for k in pob)

print(round(max_bias({"Lima": 0.5, "Cusco": 0.5}, {"Lima": 9, "Cusco": 1}), 2))""",
           ["keys faltantes"]),
        ex("S18-T2-B-E1", "S18-T2-B", "guided",
           "Para media=10, s=2, n=16, imprime el margen 1.96*s/sqrt(n) redondeado a 3.",
           "import math o numpy sqrt.",
           "margen = 1.96 * s / sqrt(n).",
           "import math\n# TODO\n",
           """import math
media, s, n = 10, 2, 16
margen = 1.96 * s / math.sqrt(n)
print("margen", round(margen, 3))""",
           ["n=0"]),
        ex("S18-T2-B-E2", "S18-T2-B", "independent",
           "Con ctrl mean=10, trat mean=13, sp=2, imprime Cohen's d redondeado a 2.",
           "d = (mb-ma)/sp.",
           "round 2.",
           "# TODO\n",
           """d = (13 - 10) / 2
print("d", round(d, 2))""",
           ["sp=0"]),
        ex("S18-T2-B-E3", "S18-T2-B", "transfer",
           "Dado diff=4, se=1.5, imprime ic95 como tupla (low, high) redondeada a 2.",
           "low = diff - 1.96*se.",
           "high = diff + 1.96*se.",
           "# TODO\n",
           """diff, se = 4, 1.5
ic = (round(diff - 1.96 * se, 2), round(diff + 1.96 * se, 2))
print("ic95", ic)""",
           ["se negativo inválido"]),
        ex("S18-T3-A-E1", "S18-T3-A", "guided",
           "Calcula pearson de x=[1,2,3,4] e y=[2,4,6,8] con np.corrcoef e imprime redondeado a 3.",
           "corrcoef[0,1].",
           "Debe ser ~1.",
           "import numpy as np\n# TODO\n",
           """import numpy as np
x = np.array([1, 2, 3, 4], dtype=float)
y = np.array([2, 4, 6, 8], dtype=float)
print("r", round(float(np.corrcoef(x, y)[0, 1]), 3))""",
           ["constante en x → nan"]),
        ex("S18-T3-A-E2", "S18-T3-A", "independent",
           "Imprime el string 'asociacion_observada' si abs(r)>0.5 para r=0.82; no uses 'causa'.",
           "if abs(r)>0.5.",
           "Mensaje fijo no causal.",
           "r = 0.82\n# TODO\n",
           """r = 0.82
print("asociacion_observada" if abs(r) > 0.5 else "asociacion_debil")""",
           ["r nan"]),
        ex("S18-T3-A-E3", "S18-T3-A", "transfer",
           "Dado z confusor, x=z, y=z (n=5 z=0..4), muestra r_raw y r tras residualizar (polyfit grado 1). Imprime ambos redondeados a 3.",
           "Residualiza x e y respecto a z.",
           "corrcoef de residuales ≈ 0 o nan controlado.",
           "import numpy as np\n# TODO\n",
           """import numpy as np
z = np.arange(5, dtype=float)
x = z.copy()
y = z.copy()
r_raw = float(np.corrcoef(x, y)[0, 1])
bx = np.polyfit(z, x, 1)
by = np.polyfit(z, y, 1)
rx = x - (bx[0] * z + bx[1])
ry = y - (by[0] * z + by[1])
# residuales ~0 → corr inestable; reporta max abs residual
print("r_raw", round(r_raw, 3))
print("max_abs_resid", round(float(max(np.max(np.abs(rx)), np.max(np.abs(ry)))), 6))""",
           ["colinealidad perfecta"]),
        ex("S18-T3-B-E1", "S18-T3-B", "guided",
           "Con montos=[10,12,11,13,50], calcula hi = Q3+1.5*IQR e imprime cuántos superan hi.",
           "quantile 0.25/0.75.",
           "suma booleana.",
           "import numpy as np\n# TODO\n",
           """import numpy as np
m = np.array([10, 12, 11, 13, 50], dtype=float)
q1, q3 = np.quantile(m, [0.25, 0.75])
iqr = q3 - q1
hi = q3 + 1.5 * iqr
print("n_hi", int((m > hi).sum()))""",
           ["sin outliers"]),
        ex("S18-T3-B-E2", "S18-T3-B", "independent",
           "Dado regiones y flags anomalia [Lima T/F, Lima T, Cusco F], imprime tasa de anomalías en Lima.",
           "Filtra región Lima.",
           "mean de flags.",
           "# TODO\n",
           """import numpy as np
region = np.array(["Lima", "Lima", "Cusco"])
flag = np.array([True, True, False])
print("tasa_Lima", float(flag[region == "Lima"].mean()))""",
           ["segmento vacío"]),
        ex("S18-T3-B-E3", "S18-T3-B", "transfer",
           "Marca anomalías Tukey en [1,2,3,4,100] e imprime lista booleana.",
           "lo = Q1-1.5IQR, hi=Q3+1.5IQR.",
           "tolist de bool.",
           "import numpy as np\n# TODO\n",
           """import numpy as np
m = np.array([1, 2, 3, 4, 100], dtype=float)
q1, q3 = np.quantile(m, [0.25, 0.75])
iqr = q3 - q1
lo, hi = q1 - 1.5 * iqr, q3 + 1.5 * iqr
print(((m < lo) | (m > hi)).tolist())""",
           ["IQR 0"]),
        ex("S18-T4-A-E1", "S18-T4-A", "guided",
           "Crea dict con claves pregunta, hipotesis, resultado(n=10, median=5.0) e imprime la pregunta.",
           "Dict literal.",
           "print evidencia['pregunta'].",
           "# TODO\n",
           """evidencia = {
    "pregunta": "¿Cuál es el ticket mediano?",
    "hipotesis": "mediana >= 5",
    "resultado": {"n": 10, "median": 5.0},
}
print(evidencia["pregunta"])""",
           ["claves faltantes"]),
        ex("S18-T4-A-E2", "S18-T4-A", "independent",
           "Dado hallazgo median=12 y umbral de decisión de negocio=15, imprime 'solo_hallazgo' si no se alcanza el umbral.",
           "Compara median < 15.",
           "No inventes decisión de campaña.",
           "median = 12\n# TODO\n",
           """median = 12
print("solo_hallazgo" if median < 15 else "candidato_decision")""",
           ["igualdad al umbral"]),
        ex("S18-T4-A-E3", "S18-T4-A", "transfer",
           "Implementa traza(pregunta, metrica, valor, limite) que imprima las 4 líneas etiquetadas.",
           "Función con 4 prints.",
           "Etiquetas fijas P/M/V/L.",
           "# TODO\n",
           """def traza(pregunta, metrica, valor, limite):
    print("P:", pregunta)
    print("M:", metrica)
    print("V:", valor)
    print("L:", limite)

traza("ticket mediano Lima", "median", 27.5, "solo web")""",
           ["None en valor"]),
        ex("S18-T4-B-E1", "S18-T4-B", "guided",
           "Con n_raw=5 y n_final=4 tras filtro, imprime data note dict con esas claves y filtros=['monto>0'].",
           "Dict con n_raw, n_final, filtros.",
           "print el dict.",
           "# TODO\n",
           """note = {"n_raw": 5, "n_final": 4, "filtros": ["monto>0"]}
print(note)""",
           ["n_final > n_raw inválido"]),
        ex("S18-T4-B-E2", "S18-T4-B", "independent",
           "Calcula sha1 hex de b'a,b\\n1,2\\n' y muestra los primeros 8 caracteres.",
           "hashlib.sha1(...).hexdigest()[:8].",
           "encode no necesario si ya es bytes.",
           "import hashlib\n# TODO\n",
           """import hashlib
print(hashlib.sha1(b"a,b\\n1,2\\n").hexdigest()[:8])""",
           ["orden de filas cambia hash"]),
        ex("S18-T4-B-E3", "S18-T4-B", "transfer",
           "Dado df de 3 filas con monto, filtra >0, genera note con n_raw, n_final y seed=42; imprime note.",
           "len antes/después.",
           "Incluye seed.",
           "import pandas as pd\n# TODO\n",
           """import pandas as pd
df = pd.DataFrame({"monto": [1.0, 0.0, 3.0]})
n_raw = len(df)
df2 = df[df["monto"] > 0]
note = {"n_raw": n_raw, "n_final": len(df2), "seed": 42}
print(note)""",
           ["todo filtrado"]),
    ]

    youdo = {
        "title": "EDA honesto para CP-N2-B (inicio)",
        "context": (
            "Eres analista en un equipo de insights en Lima. Recibes un extracto sintético de tickets "
            "(sin PII real) y debes producir un EDA que distinga hallazgo, hipótesis y decisión, con "
            "incertidumbre explícita. Esto abre el capstone **CP-N2-B**."
        ),
        "objectives": [
            "Resumir distribuciones con centro, dispersión y cuantiles",
            "Diagnosticar sesgo muestral y declarar cobertura",
            "Reportar al menos un intervalo o tamaño de efecto",
            "Evitar claims causales en correlaciones y segmentos",
            "Entregar script/notebook con data notes y seed",
        ],
        "requirements": [
            "Solo datos sintéticos o anonimizados de práctica",
            "Cada conclusión referencia un cálculo (n, métrica, código)",
            "Data note con origen, filtros, n_raw/n_final",
            "Sin secretos ni credenciales",
            "Español profesional (es-PE)",
        ],
        "starterCode": """import numpy as np
import pandas as pd

# TODO: cargar sintético, resumir, sesgo, IC/efecto, data note
rng = np.random.default_rng(18)
df = pd.DataFrame({
    "region": rng.choice(["Lima", "Arequipa", "Cusco"], size=100, p=[0.7, 0.2, 0.1]),
    "monto": rng.lognormal(3.0, 0.5, 100),
})
print(df.head())
""",
        "portfolioNote": "Artefacto de inicio CP-N2-B: EDA con incertidumbre y data notes; alimenta dashboard y reportes en S19–S21.",
        "rubric": RUBRIC_STD,
    }

    selfcheck = [
        {
            "question": "¿Qué comunica mejor un ticket “típico” con outliers fuertes?",
            "options": ["Solo la media", "Mediana (y opcionalmente IQR)", "Solo el máximo", "La moda de ids"],
            "correctIndex": 1,
            "explanation": "La mediana es robusta a colas pesadas; la media se infla con outliers.",
        },
        {
            "question": "Una correlación alta entre X e Y implica:",
            "options": ["Que X causa Y", "Asociación observada (no causal por sí sola)", "Que no hay confusores", "Que el IC es innecesario"],
            "correctIndex": 1,
            "explanation": "Correlación no implica causalidad; puede haber confusores.",
        },
        {
            "question": "¿Qué debe incluir un data note mínimo?",
            "options": ["Solo el gráfico final", "Origen, filtros, n y límites de cobertura", "La contraseña del VPN", "El prompt del LLM"],
            "correctIndex": 1,
            "explanation": "El data note hace auditable y reproducible el EDA.",
        },
        {
            "question": "El sesgo de muestra ocurre cuando:",
            "options": ["std es alta", "La muestra no representa la población de interés", "Usas mediana", "n > 30"],
            "correctIndex": 1,
            "explanation": "Selección no representativa sesga estimaciones aunque el cálculo sea correcto.",
        },
    ]

    resources = {
        "docs": [
            {"label": "NumPy statistics", "url": "https://numpy.org/doc/stable/reference/routines.statistics.html", "note": "mean, quantile, corrcoef"},
            {"label": "pandas essential basic functionality", "url": "https://pandas.pydata.org/docs/user_guide/basics.html", "note": "describe, groupby"},
        ],
        "books": [
            {"label": "Think Stats (Downey)", "note": "Estadística práctica con Python"},
            {"label": "Statistical Inference vía Data Science", "note": "Marco de incertidumbre para analistas"},
        ],
        "courses": [
            {"label": "SciPy stats overview", "url": "https://docs.scipy.org/doc/scipy/reference/stats.html", "note": "Referencia de distribuciones e IC"},
        ],
    }

    i_do = "Te demuestro el EDA de CP-N2-B inicio: resúmenes, sesgo, intervalos, correlación sin causalidad y data notes con datos sintéticos peruanos."
    we_do = "Practica 24 ejercicios (guided → independent → transfer) sobre distribuciones, incertidumbre y evidencia trazable."

    ts, log = build_section(
        "section18", meta, theories, demos, exercises, i_do, we_do, youdo, selfcheck, resources
    )
    meta_updates = {
        "id": "data-engineering",
        "index": 18,
        "title": meta["title"],
        "shortTitle": meta["shortTitle"],
        "tagline": meta["tagline"],
        "estimatedHours": 12,
        "learningOutcomes": 8,
        "youDo": youdo["title"],
        "icon": "Wrench",
        "legacy_note": "Prefect/Parquet/GE demoted; target is EDA e incertidumbre for CP-N2-B start",
        "capstone": "CP-N2-B (inicio)",
    }
    return ts, log, slugs, meta_updates, youdo["title"]




# ---------------------------------------------------------------------------
# S19 — Visualización y comunicación accesible (platform id: databases-orm)
# ---------------------------------------------------------------------------


def build_s19() -> tuple[str, dict[str, str], list[tuple[str, str]], dict, str]:
    slugs = [
        ("S19-T1-A", "question-audience-chart"),
        ("S19-T1-B", "axes-scales-encodings"),
        ("S19-T2-A", "matplotlib-seaborn"),
        ("S19-T2-B", "composition-annotations-export"),
        ("S19-T3-A", "plotly-filters-tooltips"),
        ("S19-T3-B", "state-perf-a11y-alts"),
        ("S19-T4-A", "units-source-limits"),
        ("S19-T4-B", "color-contrast-alt-no-overclaim"),
    ]
    meta = {
        "id": "databases-orm",
        "index": 19,
        "title": "Visualización y comunicación accesible",
        "shortTitle": "Viz accesible",
        "tagline": "cuatro gráficos estáticos y una vista interactiva, todos con conclusión limitada a evidencia y versión no visual equivalente",
        "estimatedHours": 12,
        "level": "Competente",
        "phase": 1,
        "icon": "Database",
        "accentColor": "bg-gradient-to-br from-blue-500 to-indigo-600",
        "jobRelevance": (
            "En equipos de analytics y reporting en Perú, una **visualización accesible y honesta** es el puente entre "
            "EDA y decisiones. Esta sección (id `databases-orm` conservado) retematiza a V3 **Visualización y "
            "comunicación accesible** para el incremento **CP-N2-B (dashboard)** con Matplotlib y tablas alternativas."
        ),
        "learningOutcomes": [
            "Elegir charts por pregunta y audiencia",
            "Diseñar ejes, escalas y encodings honestos",
            "Producir figuras Matplotlib (estilo Seaborn opcional)",
            "Anotar y exportar de forma reproducible",
            "Construir vistas interactivas lógicas (filtros/tooltips simulados)",
            "Proveer alternativas accesibles y cuidar performance",
            "Documentar unidades, fuente y limitaciones",
            "Aplicar contraste, alt text y no sobreclaim",
        ],
    }

    theories = [
        Theory(
            heading="De “Bases de Datos y ORMs” a visualización accesible (mapa)",
            paragraphs=[
                "En V3, **S19 no es el path principal de SQLAlchemy/Postgres avanzado**. Aquí construyes el **dashboard de CP-N2-B**: elección de chart, escalas honestas, figuras exportables, tooltips/filtros conceptuales y alternativas no visuales.",
                "Hilo: dataset sintético de métricas por región (Lima/Arequipa/Cusco). Cada figura tiene **conclusión acotada** y **tabla o texto equivalente**.",
                "Orden: **T1 Diseño** → **T2 Estático** → **T3 Interactivo/a11y** → **T4 Integridad visual**.",
            ],
            callout=(
                "info",
                "Contenido reubicado conceptualmente",
                "Material legado de ORM/DB de este archivo **no es el camino V3 en S19**. Target: viz accesible para CP-N2-B (dashboard).",
            ),
        ),
        Theory(
            heading="pregunta, audiencia y chart choice",
            sub="S19-T1-A",
            paragraphs=[
                "Empieza por la **pregunta** y la **audiencia** (ejecutivo vs analista). Comparar magnitudes → barras; tendencia temporal → líneas; parte de un todo → cuidado con pie (preferir barras apiladas o tabla).",
                "Una sola idea principal por chart. Si hay dos preguntas, dos charts.",
                "Registra en metadata: pregunta, chart_type, audiencia.",
            ],
            code="""spec = {
    "pregunta": "¿Qué región tiene mayor ticket mediano?",
    "audiencia": "ejecutivo",
    "chart": "bar_horizontal",
    "evita": "pie_3d",
}
print(spec)
print("ok", spec["chart"] != "pie_3d")""",
            code_title="chart_choice.py",
            callout=(
                "tip",
                "Una pregunta → un chart",
                "Si necesitas un párrafo para explicar el encoding, simplifica el chart.",
            ),
        ),
        Theory(
            heading="ejes, escalas y encodings honestos",
            sub="S19-T1-B",
            paragraphs=[
                "Ejes deben incluir **cero** en barras de magnitudes absolutas (salvo justificación). No recortes el eje Y para exagerar diferencias.",
                "Encodings: posición > longitud > color > forma. Dual axis engaña con frecuencia.",
                "Escala log solo con etiqueta explícita y justificación de órdenes de magnitud.",
            ],
            code="""vals = {"Lima": 100, "Cusco": 90}
# encoding deshonesto: baseline 85
span_honesto = max(vals.values()) - 0
span_truco = max(vals.values()) - 85
print("ratio_visual_truco", round((100-90)/span_truco, 2))
print("ratio_visual_honesto", round((100-90)/span_honesto, 2))
print("recomendacion", "baseline_0_en_barras")""",
            code_title="honest_axes.py",
            callout=(
                "danger",
                "Eje recortado",
                "Un eje Y que empieza cerca del mínimo infla diferencias percibidas.",
            ),
        ),
        Theory(
            heading="Matplotlib / Seaborn",
            sub="S19-T2-A",
            paragraphs=[
                "Matplotlib es la base reproducible (`Figure`/`Axes`). Seaborn aporta estilos y APIs estadísticas; aquí usamos Matplotlib puro (Agg) para portabilidad.",
                "Siempre: título, etiquetas de ejes, unidades, leyenda si hay series múltiples.",
                "Guarda con `bbox_inches='tight'` y dpi documentado; en demos imprimimos metadatos sin archivo binario obligatorio.",
            ],
            code="""import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

fig, ax = plt.subplots(figsize=(5, 3))
ax.bar(["Lima", "Cusco"], [28.0, 22.5], color="#1f4e79")
ax.set_ylabel("Ticket mediano (PEN)")
ax.set_title("Ticket mediano por región (sintético)")
ax.set_ylim(0, 35)
meta = {"axes": 1, "ylabel": ax.get_ylabel(), "ylim": ax.get_ylim()}
plt.close(fig)
print(meta)""",
            code_title="mpl_bar.py",
            callout=(
                "tip",
                "Backend Agg",
                "En servidores y CI usa Agg; no dependas de display interactivo.",
            ),
        ),
        Theory(
            heading="composición, annotations y exportación",
            sub="S19-T2-B",
            paragraphs=[
                "Multi-panel (`subplots`) alinea comparaciones. Anota valores clave con `ax.annotate` o `bar_label` sin saturar.",
                "Export: PNG para slides, SVG/PDF para impresión. Nombre de archivo con fecha/versión.",
                "Reproduce con seed de datos y función `build_figure()` sin estado global sucio.",
            ],
            code="""import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

fig, axes = plt.subplots(1, 2, figsize=(7, 3))
axes[0].bar(["A", "B"], [3, 4])
axes[0].set_title("Volumen")
axes[1].plot([1, 2, 3], [10, 12, 11])
axes[1].set_title("Tendencia")
for ax in axes:
    ax.set_ylim(bottom=0)
# simular export metadata
export = {"fmt": "png", "dpi": 120, "panels": 2, "name": "fig_s19_v1.png"}
plt.close(fig)
print(export)""",
            code_title="compose_export.py",
            callout=(
                "tip",
                "Función pura de figura",
                "build_figure(df) -> fig facilita tests y re-render del dashboard.",
            ),
        ),
        Theory(
            heading="Plotly / filtros / tooltips (modelo de datos)",
            sub="S19-T3-A",
            paragraphs=[
                "En entornos sin Plotly instalado, modelamos la **vista interactiva** como especificación: campos de filtro, payload de tooltip y subset de filas.",
                "Tooltips deben mostrar unidades y n, no solo el valor bonito. Filtros no deben ocultar el sesgo muestral.",
                "La conclusión del viewport filtrado debe recalcularse, no reutilizar el texto global.",
            ],
            code="""rows = [
    {"region": "Lima", "monto": 28.0, "n": 40},
    {"region": "Cusco", "monto": 22.5, "n": 32},
    {"region": "Arequipa", "monto": 24.0, "n": 28},
]
filtro = "Lima"
vista = [r for r in rows if r["region"] == filtro]
tooltip = {**vista[0], "unidad": "PEN", "nota": "sintético"}
print("filtro", filtro)
print("tooltip", tooltip)""",
            code_title="interactive_spec.py",
            callout=(
                "info",
                "Spec antes de librería",
                "Si el modelo de tooltip/filtro es claro, migrar a Plotly/Streamlit es mecánico.",
            ),
        ),
        Theory(
            heading="estado, performance y alternativas accesibles",
            sub="S19-T3-B",
            paragraphs=[
                "El **estado** del dashboard (filtros activos) debe ser serializable. Evita recalcular todo el universo en cada hover si n es grande: preagrega.",
                "Alternativa accesible: tabla ordenable + resumen textual con los mismos números que el chart.",
                "Performance: limita puntos en scatter (sample o aggregate); documenta si hay sampling.",
            ],
            code="""state = {"filtro_region": "Lima", "metric": "median"}
chart_value = 28.0
alt_table = [{"region": "Lima", "ticket_mediano_pen": 28.0, "n": 40}]
alt_text = f"En {alt_table[0]['region']}, ticket mediano {alt_table[0]['ticket_mediano_pen']} PEN (n={alt_table[0]['n']})."
print(state)
print(alt_text)
print("match", chart_value == alt_table[0]["ticket_mediano_pen"])""",
            code_title="a11y_alt.py",
            callout=(
                "success",
                "Paridad numérica",
                "La alternativa no visual debe coincidir con el chart a la misma precisión publicada.",
            ),
        ),
        Theory(
            heading="unidades, fuente y limitaciones",
            sub="S19-T4-A",
            paragraphs=[
                "Cada eje y tooltip lleva **unidad** (PEN, %, tickets). Fuente: sistema sintético / corte de fecha. Limitaciones: cobertura, sesgo, n bajo.",
                "Pie de figura: `Fuente: … | Corte: … | n=… | Limitación: …`.",
                "Sin fuente, el gráfico no entra al portfolio de CP-N2-B.",
            ],
            code="""caption = {
    "unidad": "PEN",
    "fuente": "dataset sintético curso",
    "corte": "2024-06-30",
    "n": 100,
    "limitacion": "solo canal web; no generalizar a tienda",
}
print(" | ".join(f"{k}: {v}" for k, v in caption.items()))""",
            code_title="caption.py",
            callout=(
                "warning",
                "Unidad omitida",
                "“28” sin PEN o sin % es un defecto de reporting.",
            ),
        ),
        Theory(
            heading="color, contraste, texto alternativo y no sobreclaim",
            sub="S19-T4-B",
            paragraphs=[
                "Contraste suficiente (texto/fondo). No uses solo color para codificar categorías críticas: añade patrón o etiqueta.",
                "**Alt text**: describe el hallazgo principal y n, no “imagen de barras”.",
                "No sobreclaim: “Lima lidera en la muestra web” ≠ “Lima es la mejor región del país”.",
            ],
            code="""alt = (
    "Barras del ticket mediano sintético: Lima 28 PEN (n=40), "
    "Arequipa 24 (n=28), Cusco 22.5 (n=32). Eje Y desde 0."
)
claim_ok = "En la muestra web sintética, Lima muestra el ticket mediano más alto."
claim_bad = "Lima es la región más rentable del Perú."
print("alt_len", len(alt))
print("usa_claim_ok", True)
print("evita", claim_bad[:20] + "...")""",
            code_title="alt_claim.py",
            callout=(
                "danger",
                "Sobreclaim",
                "El lenguaje del título no puede exceder la evidencia del EDA (S18).",
            ),
        ),
    ]

    demos = [
        Demo(
            "S19-T1-A-DEMO", "S19-T1-A",
            "Elegir chart alineado a pregunta ejecutiva de comparación regional",
            """pregunta = "Comparar ticket mediano entre regiones"
audiencia = "VP de operaciones"
candidates = ["bar", "line", "pie_3d", "scatter"]
score = {"bar": 3, "line": 1, "pie_3d": 0, "scatter": 1}
best = max(candidates, key=lambda c: score[c])
print({"pregunta": pregunta, "audiencia": audiencia, "chart": best})
print("rechaza_pie_3d", score["pie_3d"] == 0)""",
            "La elección de chart se documenta como decisión de diseño, no estética.",
            title="demo_chart_choice.py",
        ),
        Demo(
            "S19-T1-B-DEMO", "S19-T1-B",
            "Cuantificar distorsión de un eje Y recortado en barras",
            """a, b = 100.0, 92.0
# honesto baseline 0
perc_h = (a - b) / a
# truco baseline 90
perc_t = (a - b) / (a - 90)
print("diff_abs", a - b)
print("fraccion_altura_honesta", round(perc_h, 3))
print("fraccion_altura_truco", round(perc_t, 3))
print("factor_inflacion", round(perc_t / perc_h, 2))""",
            "Mostrar el factor de inflación educa sobre encodings deshonestos.",
            title="demo_axes.py",
        ),
        Demo(
            "S19-T2-A-DEMO", "S19-T2-A",
            "Componer barra Matplotlib con ylim desde 0 y etiqueta de unidad",
            """import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

reg = ["Lima", "Arequipa", "Cusco"]
val = [28.0, 24.0, 22.5]
fig, ax = plt.subplots()
bars = ax.bar(reg, val, color="#2c5282")
ax.set_ylabel("PEN")
ax.set_title("Ticket mediano (sintético)")
ax.set_ylim(0, max(val) * 1.2)
ax.bar_label(bars, fmt="%.1f")
print("ylim0", ax.get_ylim()[0] == 0)
print("ylabel", ax.get_ylabel())
plt.close(fig)""",
            "Figura mínima viable para el dashboard estático.",
            title="demo_mpl.py",
        ),
        Demo(
            "S19-T2-B-DEMO", "S19-T2-B",
            "Anotar y exportar metadata reproducible de figura multi-panel",
            """import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

fig, (ax0, ax1) = plt.subplots(1, 2, figsize=(8, 3))
ax0.bar(["Lima", "Cusco"], [40, 32])
ax0.set_title("n por región")
ax0.set_ylim(0, 50)
ax1.barh(["Lima", "Cusco"], [28, 22.5])
ax1.set_title("mediana PEN")
ax1.set_xlim(0, 35)
export = {"file": "cp_n2b_dashboard_v1.png", "dpi": 120, "panels": 2, "seed_data": 19}
plt.close(fig)
print(export)""",
            "Composición + nombre versionado habilita re-render del portfolio.",
            title="demo_compose.py",
        ),
        Demo(
            "S19-T3-A-DEMO", "S19-T3-A",
            "Vista interactiva lógica con filtro y tooltip honesto",
            """data = [
    {"region": "Lima", "median": 28.0, "n": 40},
    {"region": "Cusco", "median": 22.5, "n": 32},
]
def view(region):
    row = next(r for r in data if r["region"] == region)
    return {
        "tooltip": f"{row['region']}: {row['median']} PEN (n={row['n']})",
        "filtro": region,
        "unidad": "PEN",
    }
print(view("Lima"))
print(view("Cusco")["tooltip"])""",
            "Tooltip con unidad y n evita lecturas superficiales.",
            title="demo_tooltip.py",
        ),
        Demo(
            "S19-T3-B-DEMO", "S19-T3-B",
            "Ofrecer alternativa tabular/textual con paridad al chart",
            """chart = {"Lima": 28.0, "Cusco": 22.5}
table = [{"region": k, "ticket_mediano_pen": v} for k, v in chart.items()]
text = "; ".join(f"{r['region']}={r['ticket_mediano_pen']} PEN" for r in table)
print(table)
print(text)
print("parity", all(chart[r["region"]] == r["ticket_mediano_pen"] for r in table))""",
            "La versión no visual es requisito del gate de dashboard accesible.",
            title="demo_a11y.py",
        ),
        Demo(
            "S19-T4-A-DEMO", "S19-T4-A",
            "Etiquetar unidades, fuente y limitaciones en caption estructurado",
            """cap = {
    "titulo": "Ticket mediano por región",
    "unidad": "PEN",
    "fuente": "sintético CP-N2-B",
    "corte": "2024-06-30",
    "limitacion": "canal web; n bajo en Cusco",
}
print("pie", f"Unidad: {cap['unidad']} | Fuente: {cap['fuente']} | Corte: {cap['corte']} | Límite: {cap['limitacion']}")""",
            "Caption completo es parte del entregable, no un extra.",
            title="demo_caption.py",
        ),
        Demo(
            "S19-T4-B-DEMO", "S19-T4-B",
            "Validar alt text y rechazar sobreclaim causal/nacional",
            """alt = "Barras: Lima 28 PEN, Arequipa 24, Cusco 22.5; muestra web sintética n=100."
claims = [
    ("Lima lidera el ticket mediano en la muestra web", True),
    ("Lima es la mejor región del Perú", False),
]
for c, ok in claims:
    print(c[:40], "=>", "PERMITIDO" if ok else "RECHAZADO")
print("alt_words", len(alt.split()))""",
            "Contraste de claims entrena el lenguaje del dashboard.",
            title="demo_claims.py",
        ),
    ]

    exercises = [
        ex("S19-T1-A-E1", "S19-T1-A", "guided",
           "Dada pregunta de comparación entre 3 regiones, imprime chart recomendado 'bar'.",
           "Asigna chart='bar'.",
           "print el valor.",
           "pregunta = 'comparar regiones'\n# TODO\n",
           """pregunta = "comparar regiones"
chart = "bar"
print(chart)""",
           ["serie temporal mal clasificada"]),
        ex("S19-T1-A-E2", "S19-T1-A", "independent",
           "Devuelve dict pregunta/audiencia/chart para audiencia ejecutivo y comparación de totales.",
           "Incluye las 3 claves.",
           "chart bar.",
           "# TODO\n",
           """print({"pregunta": "totales por región", "audiencia": "ejecutivo", "chart": "bar"})""",
           ["audiencia técnica puede preferir table"]),
        ex("S19-T1-A-E3", "S19-T1-A", "transfer",
           "Escribe elige_chart(pregunta) que devuelva 'line' si 'tendencia' in pregunta else 'bar'; prueba con dos strings.",
           "in para substring.",
           "Dos prints.",
           "# TODO\n",
           """def elige_chart(pregunta):
    return "line" if "tendencia" in pregunta.lower() else "bar"
print(elige_chart("tendencia mensual"))
print(elige_chart("comparar regiones"))""",
           ["mayúsculas"]),
        ex("S19-T1-B-E1", "S19-T1-B", "guided",
           "Con valores 50 y 45, imprime inflación visual si baseline=40 vs 0 (ratio de alturas de la diferencia).",
           "altura_truco=(50-45)/(50-40); altura_h=(50-45)/50.",
           "factor = truco/honesto.",
           "# TODO\n",
           """truco = (50 - 45) / (50 - 40)
hon = (50 - 45) / 50
print("factor", round(truco / hon, 2))""",
           ["baseline > min"]),
        ex("S19-T1-B-E2", "S19-T1-B", "independent",
           "Imprime 'honesto' si ylim_bottom==0 else 'revisar'.",
           "Variable ylim_bottom=0.",
           "print condicional.",
           "ylim_bottom = 0\n# TODO\n",
           """ylim_bottom = 0
print("honesto" if ylim_bottom == 0 else "revisar")""",
           ["líneas de índice pueden no empezar en 0"]),
        ex("S19-T1-B-E3", "S19-T1-B", "transfer",
           "Dado encoding='dual_axis', imprime 'riesgo_alto' ; si 'position', 'ok'.",
           "if/elif.",
           "Prueba dual_axis.",
           "encoding = 'dual_axis'\n# TODO\n",
           """encoding = "dual_axis"
print("riesgo_alto" if encoding == "dual_axis" else "ok")""",
           ["color-only"]),
        ex("S19-T2-A-E1", "S19-T2-A", "guided",
           "Crea figura bar de [1,2] y verifica get_ylim()[0]==0 tras set_ylim(0,3); imprime True/False.",
           "matplotlib Agg.",
           "plt.close.",
           "import matplotlib\nmatplotlib.use('Agg')\nimport matplotlib.pyplot as plt\n# TODO\n",
           """import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
fig, ax = plt.subplots()
ax.bar(["a", "b"], [1, 2])
ax.set_ylim(0, 3)
print(ax.get_ylim()[0] == 0)
plt.close(fig)""",
           ["olvidar close"]),
        ex("S19-T2-A-E2", "S19-T2-A", "independent",
           "Asigna ylabel 'PEN' a un Axes y print get_ylabel().",
           "set_ylabel.",
           "Agg backend.",
           "import matplotlib\nmatplotlib.use('Agg')\nimport matplotlib.pyplot as plt\n# TODO\n",
           """import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
fig, ax = plt.subplots()
ax.set_ylabel("PEN")
print(ax.get_ylabel())
plt.close(fig)""",
           ["ylabel vacío"]),
        ex("S19-T2-A-E3", "S19-T2-A", "transfer",
           "Función meta_bar(labels, values) crea bar chart y devuelve dict n_bars y ylim0; print con ['A','B'] [3,4].",
           "len(values).",
           "set_ylim(0, max*1.2).",
           "import matplotlib\nmatplotlib.use('Agg')\nimport matplotlib.pyplot as plt\n# TODO\n",
           """import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

def meta_bar(labels, values):
    fig, ax = plt.subplots()
    ax.bar(labels, values)
    ax.set_ylim(0, max(values) * 1.2)
    out = {"n_bars": len(values), "ylim0": ax.get_ylim()[0]}
    plt.close(fig)
    return out
print(meta_bar(["A", "B"], [3, 4]))""",
           ["values vacíos"]),
        ex("S19-T2-B-E1", "S19-T2-B", "guided",
           "Imprime export dict con fmt png, dpi 120, panels 2.",
           "Dict literal.",
           "print.",
           "# TODO\n",
           """print({"fmt": "png", "dpi": 120, "panels": 2})""",
           ["dpi 0"]),
        ex("S19-T2-B-E2", "S19-T2-B", "independent",
           "Construye nombre fig_cpn2b_v{version}.png para version=3.",
           "f-string.",
           "print nombre.",
           "version = 3\n# TODO\n",
           """version = 3
print(f"fig_cpn2b_v{version}.png")""",
           ["version string"]),
        ex("S19-T2-B-E3", "S19-T2-B", "transfer",
           "Crea 1x2 subplots, set titles Vol y Med, devuelve lista de titles; print.",
           "ax.get_title().",
           "close fig.",
           "import matplotlib\nmatplotlib.use('Agg')\nimport matplotlib.pyplot as plt\n# TODO\n",
           """import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
fig, axes = plt.subplots(1, 2)
axes[0].set_title("Vol")
axes[1].set_title("Med")
print([ax.get_title() for ax in axes])
plt.close(fig)""",
           ["orientación 2x1"]),
        ex("S19-T3-A-E1", "S19-T3-A", "guided",
           "Filtra rows por region=='Lima' e imprime el median del primero.",
           "list comp o next.",
           "print valor.",
           "rows=[{'region':'Lima','median':28},{'region':'Cusco','median':22}]\n# TODO\n",
           """rows = [{"region": "Lima", "median": 28}, {"region": "Cusco", "median": 22}]
print(next(r for r in rows if r["region"] == "Lima")["median"])""",
           ["sin match"]),
        ex("S19-T3-A-E2", "S19-T3-A", "independent",
           "Arma tooltip string '{region}: {v} PEN (n={n})' con Lima 28 n=40.",
           "f-string.",
           "print.",
           "# TODO\n",
           """print(f"Lima: {28} PEN (n={40})")""",
           ["sin unidad"]),
        ex("S19-T3-A-E3", "S19-T3-A", "transfer",
           "Implementa tooltip(row) y pruébalo con dict region/median/n.",
           "Incluye unidad PEN.",
           "Función pura.",
           "# TODO\n",
           """def tooltip(row):
    return f"{row['region']}: {row['median']} PEN (n={row['n']})"
print(tooltip({"region": "Cusco", "median": 22.5, "n": 32}))""",
           ["keys faltantes"]),
        ex("S19-T3-B-E1", "S19-T3-B", "guided",
           "Dado chart={'Lima':1} y table=[{'region':'Lima','v':1}], imprime parity True si coinciden.",
           "Compara valores.",
           "print True.",
           "# TODO\n",
           """chart = {"Lima": 1}
table = [{"region": "Lima", "v": 1}]
print(chart["Lima"] == table[0]["v"])""",
           ["float redondeo"]),
        ex("S19-T3-B-E2", "S19-T3-B", "independent",
           "Serializa state filtro_region=Lima a string JSON compacto (json.dumps).",
           "import json.",
           "print.",
           "import json\n# TODO\n",
           """import json
print(json.dumps({"filtro_region": "Lima"}, ensure_ascii=False))""",
           ["estado no serializable"]),
        ex("S19-T3-B-E3", "S19-T3-B", "transfer",
           "Genera alt_text desde table list de dicts region/v; une con '; '.",
           "join.",
           "Incluye PEN.",
           "table=[{'region':'Lima','v':28},{'region':'Cusco','v':22}]\n# TODO\n",
           """table = [{"region": "Lima", "v": 28}, {"region": "Cusco", "v": 22}]
print("; ".join(f"{r['region']}={r['v']} PEN" for r in table))""",
           ["tabla vacía"]),
        ex("S19-T4-A-E1", "S19-T4-A", "guided",
           "Imprime caption con unidad PEN y fuente sintetico.",
           "f-string o dict.",
           "Ambas menciones.",
           "# TODO\n",
           """print("unidad=PEN | fuente=sintetico")""",
           ["fuente vacía"]),
        ex("S19-T4-A-E2", "S19-T4-A", "independent",
           "Valida que caption dict tenga claves unidad, fuente, limitacion; imprime True.",
           "set de keys.",
           "issuperset.",
           "cap={'unidad':'PEN','fuente':'x','limitacion':'web'}\n# TODO\n",
           """cap = {"unidad": "PEN", "fuente": "x", "limitacion": "web"}
print(set(cap) >= {"unidad", "fuente", "limitacion"})""",
           ["typo en clave"]),
        ex("S19-T4-A-E3", "S19-T4-A", "transfer",
           "Función pie(cap) devuelve string unido con ' | ' de items; prueba con 2 claves.",
           "join items.",
           "orden de inserción dict.",
           "# TODO\n",
           """def pie(cap):
    return " | ".join(f"{k}: {v}" for k, v in cap.items())
print(pie({"unidad": "PEN", "n": 10}))""",
           ["valores None"]),
        ex("S19-T4-B-E1", "S19-T4-B", "guided",
           "Si claim contiene 'del Perú' sin 'muestra', imprime RECHAZADO.",
           "in checks.",
           "Caso: 'mejor del Perú'.",
           "claim = 'Lima es la mejor del Perú'\n# TODO\n",
           """claim = "Lima es la mejor del Perú"
print("RECHAZADO" if ("del Perú" in claim and "muestra" not in claim) else "OK")""",
           ["claims legítimos locales"]),
        ex("S19-T4-B-E2", "S19-T4-B", "independent",
           "Alt text debe mencionar 'n='; imprime True si 'n=' in alt.",
           "substring.",
           "alt de ejemplo.",
           "alt = 'Lima 28 PEN n=40'\n# TODO\n",
           """alt = "Lima 28 PEN n=40"
print("n=" in alt or "n=" in alt.replace("n=", "n="))
print("n=" in "Lima 28 PEN n=40")""",
           ["n sin valor"]),
        ex("S19-T4-B-E3", "S19-T4-B", "transfer",
           "classify_claim(text) -> PERMITIDO si 'muestra' in text else RECHAZADO; prueba 2 textos.",
           "Función binaria simple didáctica.",
           "Dos prints.",
           "# TODO\n",
           """def classify_claim(text):
    return "PERMITIDO" if "muestra" in text else "RECHAZADO"
print(classify_claim("lidera en la muestra web"))
print(classify_claim("es la mejor del país"))""",
           ["falsos positivos"]),
    ]

    youdo = {
        "title": "Dashboard accesible CP-N2-B",
        "context": (
            "Construye el incremento dashboard de **CP-N2-B**: al menos cuatro gráficos estáticos y una vista "
            "interactiva lógica, cada uno con conclusión limitada a evidencia y alternativa no visual."
        ),
        "objectives": [
            "Elegir charts por pregunta/audiencia",
            "Ejes honestos y unidades visibles",
            "Export versionado + captions",
            "Alt text y sin sobreclaim",
            "Paridad numérica chart/tabla",
        ],
        "requirements": [
            "Datos sintéticos únicamente",
            "ylim de barras desde 0 salvo justificación escrita",
            "Caption con fuente y limitación",
            "Alt text por figura",
            "es-PE en títulos y conclusiones",
        ],
        "starterCode": """import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd

df = pd.DataFrame({
    "region": ["Lima", "Arequipa", "Cusco"],
    "median_pen": [28.0, 24.0, 22.5],
    "n": [40, 28, 32],
})
# TODO: 4 figuras + alt/caption + tabla equivalente
print(df)
""",
        "portfolioNote": "Dashboard del factory CP-N2-B; se integra con Excel (S20) y reportes (S21).",
        "rubric": RUBRIC_STD,
    }

    selfcheck = [
        {
            "question": "Para comparar magnitudes entre categorías, ¿qué chart es usualmente preferible?",
            "options": ["Pie 3D", "Barras con baseline 0", "Dual axis sin escala", "Word cloud"],
            "correctIndex": 1,
            "explanation": "Las barras con baseline 0 comunican magnitudes de forma honesta.",
        },
        {
            "question": "Una alternativa accesible debe:",
            "options": ["Ser opcional siempre", "Repetir los mismos números clave del chart", "Solo ser una imagen más grande", "Eliminar las unidades"],
            "correctIndex": 1,
            "explanation": "Paridad numérica entre chart y tabla/texto.",
        },
        {
            "question": "“Lima es la mejor región del Perú” a partir de una muestra web es:",
            "options": ["Un claim permitido", "Sobreclaim / generalización indebida", "Un alt text correcto", "Una unidad"],
            "correctIndex": 1,
            "explanation": "El lenguaje no debe exceder la cobertura de la muestra.",
        },
        {
            "question": "El caption de un gráfico de portfolio debe incluir:",
            "options": ["Solo el color favorito", "Unidad, fuente y limitaciones", "La contraseña del BI", "Nada"],
            "correctIndex": 1,
            "explanation": "Trazabilidad y honestidad metodológica.",
        },
    ]

    resources = {
        "docs": [
            {"label": "Matplotlib tutorials", "url": "https://matplotlib.org/stable/tutorials/index.html", "note": "Figuras y exportación"},
            {"label": "Data Visualization Society resources", "url": "https://www.datavisualizationsociety.org/", "note": "Comunidad y ética viz"},
        ],
        "books": [
            {"label": "Fundamentals of Data Visualization (Wilke)", "note": "Encodings y honestidad visual"},
            {"label": "Storytelling with Data (Knaflic)", "note": "Audiencia y claridad"},
        ],
        "courses": [
            {"label": "Matplotlib cheatsheets", "url": "https://matplotlib.org/cheatsheets/", "note": "Referencia rápida"},
        ],
    }

    i_do = "Te muestro cómo diseñar charts honestos, exportables y con alternativa accesible para el dashboard CP-N2-B."
    we_do = "24 ejercicios de elección de chart, ejes, Matplotlib, tooltips lógicos, a11y y claims."
    ts, log = build_section(
        "section19", meta, theories, demos, exercises, i_do, we_do, youdo, selfcheck, resources
    )
    meta_updates = {
        "id": "databases-orm",
        "index": 19,
        "title": meta["title"],
        "shortTitle": meta["shortTitle"],
        "tagline": meta["tagline"],
        "estimatedHours": 12,
        "learningOutcomes": 8,
        "youDo": youdo["title"],
        "icon": "Database",
        "legacy_note": "ORM/DB demoted; target is accessible visualization for CP-N2-B dashboard",
        "capstone": "CP-N2-B (dashboard)",
    }
    return ts, log, slugs, meta_updates, youdo["title"]



# ---------------------------------------------------------------------------
# S20 — Automatización robusta de Excel (platform id: rag)
# ---------------------------------------------------------------------------


def build_s20() -> tuple[str, dict[str, str], list[tuple[str, str]], dict, str]:
    slugs = [
        ("S20-T1-A", "sheets-cells-tables-named"),
        ("S20-T1-B", "formulas-vs-cached-values"),
        ("S20-T2-A", "styles-charts-templates"),
        ("S20-T2-B", "dates-locales-merged-protect"),
        ("S20-T3-A", "reconcile-pivots"),
        ("S20-T3-B", "validation-structure-preserve"),
        ("S20-T4-A", "batch-corrupt-locks"),
        ("S20-T4-B", "backups-idempotency-visual-tests"),
    ]
    meta = {
        "id": "rag",
        "index": 20,
        "title": "Automatización robusta de Excel",
        "shortTitle": "Excel factory",
        "tagline": "adaptador que lee los formatos sintéticos del VP, produce un workbook de resultados sin dañar la plantilla y deja manifest de cambios",
        "estimatedHours": 12,
        "level": "Competente",
        "phase": 1,
        "icon": "MessageSquare",
        "accentColor": "bg-gradient-to-br from-blue-500 to-indigo-600",
        "jobRelevance": (
            "En finanzas, operaciones y reporting en Perú, **Excel sigue siendo el contrato de entrega**. Esta sección "
            "(id `rag` conservado) retematiza a V3 **Automatización robusta de Excel** para **CP-N2-B (excel factory)** "
            "con openpyxl, plantillas, conciliación y manifests — sin PII real."
        ),
        "learningOutcomes": [
            "Manipular sheets, celdas, tablas y named ranges",
            "Distinguir fórmulas de valores cacheados",
            "Aplicar estilos, charts y plantillas",
            "Manejar fechas, locales, merged cells y protección",
            "Conciliar totales y trabajar con pivots lógicos",
            "Preservar estructura y validaciones",
            "Operar batch con corruptos y locks",
            "Garantizar backups, idempotencia y tests estructurales",
        ],
    }

    theories = [
        Theory(
            heading="De “RAG en producción” a Excel factory (mapa)",
            paragraphs=[
                "En V3, **S20 no es el path principal de embeddings/vector stores**. Aquí construyes el **excel factory de CP-N2-B**: leer/escribir celdas, no pisar plantillas, conciliar totales y dejar **manifest** de cambios.",
                "Usamos **openpyxl** en memoria/temp. Datos sintéticos de regiones y montos.",
                "Orden: **T1 Modelo de libro** → **T2 Presentación** → **T3 Conciliación** → **T4 Operación robusta**.",
            ],
            callout=(
                "info",
                "Contenido reubicado conceptualmente",
                "Material legado de RAG de este archivo **no es el camino V3 en S20**. Target: automatización Excel para CP-N2-B.",
            ),
        ),
        Theory(
            heading="sheets, celdas, tablas y named ranges",
            sub="S20-T1-A",
            paragraphs=[
                "Un workbook tiene **sheets**; las celdas se direccionan por coordenada (`A1`) o índice. Las **tablas** (display name) y **defined names** dan estabilidad frente a inserts de filas.",
                "Lee por nombre de hoja, no por posición frágil si el VP reordena.",
                "Documenta el mapa: hoja → rango → significado.",
            ],
            code="""from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Entrada"
ws["A1"] = "region"
ws["B1"] = "monto"
ws["A2"] = "Lima"
ws["B2"] = 28.0
wb.create_sheet("Salida")
print(wb.sheetnames)
print(ws["A2"].value, ws["B2"].value)""",
            code_title="sheets_cells.py",
            callout=(
                "tip",
                "Nombres estables",
                "Prefiere títulos de hoja y headers explícitos a “la segunda columna del primer sheet”.",
            ),
        ),
        Theory(
            heading="fórmulas vs valores cacheados",
            sub="S20-T1-B",
            paragraphs=[
                "openpyxl por defecto no evalúa fórmulas de Excel: lees el **string de fórmula** o un **valor cacheado** si existe en el archivo.",
                "Para pipelines Python, suele ser más seguro **calcular en Python** y escribir valores, o documentar dependencia de Excel para recalcular.",
                "Nunca asumas que `cell.value` numérico implica ausencia de fórmula en origen.",
            ],
            code="""from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws["A1"] = 10
ws["A2"] = 5
ws["A3"] = "=A1+A2"
print("formula", ws["A3"].value)
# cálculo python equivalente
print("python_sum", ws["A1"].value + ws["A2"].value)""",
            code_title="formulas.py",
            callout=(
                "warning",
                "Sin motor Excel",
                "En CI Linux, no hay Excel: no dependas de valores cacheados no controlados.",
            ),
        ),
        Theory(
            heading="estilos, charts y plantillas",
            sub="S20-T2-A",
            paragraphs=[
                "Las **plantillas** (.xlsx) del VP traen estilos corporativos: no reconstruyas el libro desde cero si puedes rellenar celdas marcadas.",
                "Estilos (Font, PatternFill, Alignment) y charts openpyxl se aplican con cuidado para no romper named styles.",
                "Separa hoja de datos de hoja de presentación.",
            ],
            code="""from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws["A1"] = "KPI"
ws["A1"].font = Font(bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws["B1"] = 28.0
print(ws["A1"].font.bold, ws["B1"].value)""",
            code_title="styles.py",
            callout=(
                "tip",
                "Plantilla intocable",
                "Copia la plantilla a un output path; nunca escribas sobre el master sin backup.",
            ),
        ),
        Theory(
            heading="fechas, locales, celdas combinadas y protección",
            sub="S20-T2-B",
            paragraphs=[
                "Las **fechas** en Excel son seriales; openpyxl puede devolver `datetime`. Locales de display (dd/mm vs mm/dd) son del cliente Excel, no del valor.",
                "**Merged cells**: escribe en la celda ancla (top-left). Leer celdas no ancla puede devolver None.",
                "Protección de hoja/celda es señal al usuario, no seguridad criptográfica.",
            ],
            code="""from openpyxl import Workbook
from datetime import date

wb = Workbook()
ws = wb.active
ws["A1"] = date(2024, 6, 30)
ws.merge_cells("B1:C1")
ws["B1"] = "Lima-Arequipa"
print(ws["A1"].value.isoformat())
print("merge", ws.merged_cells.ranges)
print("C1_is_none", ws["C1"].value)""",
            code_title="dates_merge.py",
            callout=(
                "warning",
                "Merged traps",
                "Automatizar sin mapear merges rompe layouts del VP.",
            ),
        ),
        Theory(
            heading="conciliación y pivots",
            sub="S20-T3-A",
            paragraphs=[
                "Antes de entregar, **concilia**: suma de detalle = total de portada; conteos por región = n del EDA.",
                "Los pivots de Excel no siempre se refrescan con openpyxl: calcula la tabla pivote en pandas y escríbela como valores.",
                "Tolerancia monetaria (p. ej. 0.01 PEN) debe documentarse.",
            ],
            code="""import pandas as pd

det = pd.DataFrame({"region": ["Lima", "Lima", "Cusco"], "monto": [10.0, 5.0, 7.0]})
tot_portada = 22.0
tot_det = float(det["monto"].sum())
pivot = det.groupby("region", as_index=False)["monto"].sum()
print(pivot.to_dict(orient="list"))
print("ok", abs(tot_det - tot_portada) < 0.01)""",
            code_title="reconcile.py",
            callout=(
                "success",
                "Gate de conciliación",
                "Si falla la conciliación, no se emite el workbook final.",
            ),
        ),
        Theory(
            heading="reglas de validación y preservación de estructura",
            sub="S20-T3-B",
            paragraphs=[
                "Data validation (listas, rangos) en plantillas debe **preservarse**. Al escribir, no borres rangos enteros si puedes actualizar celdas puntuales.",
                "Valida headers esperados antes de cargar filas.",
                "Estructura: mismas columnas, orden y tipos contractuales del VP.",
            ],
            code="""expected = ["region", "monto", "n"]
headers = ["region", "monto", "n"]
print("structure_ok", headers == expected)
# validación de dominio
regiones = {"Lima", "Arequipa", "Cusco"}
row = {"region": "Lima", "monto": 10.0}
print("domain_ok", row["region"] in regiones)""",
            code_title="structure.py",
            callout=(
                "tip",
                "Fail fast en headers",
                "Si el header no coincide, aborta con mensaje claro al manifest.",
            ),
        ),
        Theory(
            heading="batch, archivos corruptos y locks",
            sub="S20-T4-A",
            paragraphs=[
                "Procesa carpetas en batch con try/except por archivo. Corruptos → cuarentena + log; no detengas todo el lote sin política.",
                "Locks (`~$' file`, PermissionError) se reintentan o se reportan.",
                "Manifest lista: ok / corrupt / locked / skipped.",
            ],
            code="""files = ["a.xlsx", "b.xlsx", "c.xlsx"]
status = {}
for f in files:
    if f.startswith("b"):
        status[f] = "corrupt"
    else:
        status[f] = "ok"
print(status)
print("ok_count", sum(v == "ok" for v in status.values()))""",
            code_title="batch.py",
            callout=(
                "info",
                "Cuarentena",
                "Mueve corruptos a /quarantine y deja hash/nombre en el log.",
            ),
        ),
        Theory(
            heading="backups, idempotencia y pruebas estructurales",
            sub="S20-T4-B",
            paragraphs=[
                "**Backup** del output anterior antes de sobrescribir. **Idempotencia**: re-ejecutar con mismos inputs produce mismo resultado lógico (o versiona outputs).",
                "Tests estructurales: sheetnames, headers, n filas, total conciliado, no fórmulas rotas esperadas.",
                "El manifest JSON es el artefacto de auditoría del excel factory.",
            ],
            code="""import json
import hashlib

payload = b"region,monto\\nLima,10\\n"
manifest = {
    "input_sha1_8": hashlib.sha1(payload).hexdigest()[:8],
    "sheets": ["Entrada", "Salida"],
    "reconcile_ok": True,
    "backup": "out/prev_results.xlsx.bak",
    "idempotent": True,
}
print(json.dumps(manifest, ensure_ascii=False))""",
            code_title="manifest.py",
            callout=(
                "success",
                "Manifest obligatorio",
                "Sin manifest, el incremento excel factory de CP-N2-B no cierra.",
            ),
        ),
    ]

    demos = [
        Demo(
            "S20-T1-A-DEMO", "S20-T1-A",
            "Leer/escribir sheets y celdas con openpyxl en workbook sintético",
            """from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Datos"
ws.append(["region", "monto"])
ws.append(["Lima", 28.0])
ws.append(["Cusco", 22.5])
out = wb.create_sheet("Resultados")
out["A1"] = "n_filas"
out["B1"] = ws.max_row - 1
print(wb.sheetnames)
print("n", out["B1"].value)
print("A2", ws["A2"].value)""",
            "Mapa de hojas estable es el primer contrato del adaptador.",
            title="demo_sheets.py",
        ),
        Demo(
            "S20-T1-B-DEMO", "S20-T1-B",
            "Distinguir fórmula almacenada vs suma calculada en Python",
            """from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws["B1"] = 10
ws["B2"] = 15
ws["B3"] = "=B1+B2"
formula = ws["B3"].value
py_val = ws["B1"].value + ws["B2"].value
print("es_formula", isinstance(formula, str) and formula.startswith("="))
print("python_value", py_val)
print("no_evaluado_por_openpyxl", formula != py_val)""",
            "El factory prefiere valores Python auditables en hojas de salida.",
            title="demo_formula.py",
        ),
        Demo(
            "S20-T2-A-DEMO", "S20-T2-A",
            "Aplicar estilo de encabezado tipo plantilla corporativa",
            """from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws["A1"] = "Region"
ws["B1"] = "Monto PEN"
for col in ("A", "B"):
    c = ws[f"{col}1"]
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center")
ws["A2"] = "Lima"
ws["B2"] = 28.0
print(ws["A1"].font.bold, ws["A1"].fill.fgColor.rgb)
print(ws["B2"].value)""",
            "Estilos consistentes sin rehacer la plantilla completa.",
            title="demo_styles.py",
        ),
        Demo(
            "S20-T2-B-DEMO", "S20-T2-B",
            "Escribir fecha ISO y respetar celda combinada (ancla)",
            """from openpyxl import Workbook
from datetime import date

wb = Workbook()
ws = wb.active
ws["A1"] = date(2024, 6, 30)
ws.merge_cells("B1:D1")
ws["B1"] = "Cobertura: Lima|Arequipa|Cusco"
print(ws["A1"].value.isoformat())
print("anchor", ws["B1"].value)
print("non_anchor_D1", ws["D1"].value)""",
            "Fechas y merges sin romper el layout del VP.",
            title="demo_dates_merge.py",
        ),
        Demo(
            "S20-T3-A-DEMO", "S20-T3-A",
            "Conciliar total de portada con detalle y pivot por región",
            """import pandas as pd

det = pd.DataFrame({
    "region": ["Lima", "Lima", "Arequipa", "Cusco"],
    "monto": [10.0, 12.0, 8.0, 5.5],
})
portada = 35.5
pivot = det.groupby("region")["monto"].sum().to_dict()
print(pivot)
print("reconcile", abs(det["monto"].sum() - portada) < 0.01)""",
            "Conciliación es el control de calidad del workbook de resultados.",
            title="demo_reconcile.py",
        ),
        Demo(
            "S20-T3-B-DEMO", "S20-T3-B",
            "Validar headers y dominio de región antes de escribir salida",
            """expected = ["region", "monto", "n"]
got = ["region", "monto", "n"]
allowed = {"Lima", "Arequipa", "Cusco"}
rows = [{"region": "Lima", "monto": 1.0, "n": 1}, {"region": "Piura", "monto": 1.0, "n": 1}]
print("headers_ok", got == expected)
bad = [r for r in rows if r["region"] not in allowed]
print("bad_regions", [r["region"] for r in bad])
print("abort", len(bad) > 0)""",
            "Fail fast preserva la estructura contractual del VP.",
            title="demo_validate.py",
        ),
        Demo(
            "S20-T4-A-DEMO", "S20-T4-A",
            "Simular batch con detección de corruptos y locks",
            """batch = [
    ("ok1.xlsx", "ok"),
    ("bad.xlsx", "corrupt"),
    ("lock.xlsx", "locked"),
    ("ok2.xlsx", "ok"),
]
manifest = {}
for name, st in batch:
    manifest[name] = st
print(manifest)
print({k: sum(1 for v in manifest.values() if v == k) for k in ("ok", "corrupt", "locked")})""",
            "El lote continúa; los fallos quedan auditados.",
            title="demo_batch.py",
        ),
        Demo(
            "S20-T4-B-DEMO", "S20-T4-B",
            "Escritura idempotente con backup lógico y tests estructurales",
            """import hashlib, json

def build_output(rows):
    # determinista
    lines = ["region,monto"] + [f"{r},{m}" for r, m in sorted(rows)]
    return "\\n".join(lines) + "\\n"

rows = [("Lima", 10), ("Cusco", 5)]
o1 = build_output(rows)
o2 = build_output(list(reversed(rows)))
manifest = {
    "sha1_8": hashlib.sha1(o1.encode()).hexdigest()[:8],
    "idempotent": o1 == o2,
    "backup": "results.prev.xlsx",
    "tests": {"has_header": o1.startswith("region,monto"), "n_data": 2},
}
print(json.dumps(manifest, ensure_ascii=False))""",
            "Idempotencia + backup + tests cierran el excel factory.",
            title="demo_idempotent.py",
        ),
    ]

    exercises = [
        ex("S20-T1-A-E1", "S20-T1-A", "guided",
           "Crea Workbook, pon título de hoja 'Entrada' y A1='region'; imprime sheetnames y A1.",
           "ws.title y ws['A1'].",
           "from openpyxl import Workbook.",
           "from openpyxl import Workbook\n# TODO\n",
           """from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "Entrada"
ws["A1"] = "region"
print(wb.sheetnames)
print(ws["A1"].value)""",
           ["nombre con espacios"]),
        ex("S20-T1-A-E2", "S20-T1-A", "independent",
           "append header y una fila Lima/10.0; imprime max_row.",
           "ws.append.",
           "max_row incluye header.",
           "from openpyxl import Workbook\n# TODO\n",
           """from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.append(["region", "monto"])
ws.append(["Lima", 10.0])
print(ws.max_row)""",
           ["filas vacías"]),
        ex("S20-T1-A-E3", "S20-T1-A", "transfer",
           "Crea hojas Datos y Salida; imprime sheetnames ordenados por creación.",
           "create_sheet.",
           "Renombra active primero.",
           "from openpyxl import Workbook\n# TODO\n",
           """from openpyxl import Workbook
wb = Workbook()
wb.active.title = "Datos"
wb.create_sheet("Salida")
print(wb.sheetnames)""",
           ["duplicar nombre"]),
        ex("S20-T1-B-E1", "S20-T1-B", "guided",
           "Asigna fórmula =A1+A2 y print si startswith('=').",
           "cell value string.",
           "True esperado.",
           "from openpyxl import Workbook\n# TODO\n",
           """from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws["A3"] = "=A1+A2"
print(str(ws["A3"].value).startswith("="))""",
           ["valor cacheado"]),
        ex("S20-T1-B-E2", "S20-T1-B", "independent",
           "Con A1=3 A2=4 calcula suma python e imprime 7 sin evaluar fórmula Excel.",
           "Suma .value numéricos.",
           "No uses data_only aquí.",
           "from openpyxl import Workbook\n# TODO\n",
           """from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws["A1"] = 3
ws["A2"] = 4
print(ws["A1"].value + ws["A2"].value)""",
           ["None en celdas"]),
        ex("S20-T1-B-E3", "S20-T1-B", "transfer",
           "Función es_formula(v) True si str y empieza con =; prueba con '=A1' y 3.",
           "isinstance str.",
           "Dos prints.",
           "# TODO\n",
           """def es_formula(v):
    return isinstance(v, str) and v.startswith("=")
print(es_formula("=A1"))
print(es_formula(3))""",
           ["espacios antes de ="]),
        ex("S20-T2-A-E1", "S20-T2-A", "guided",
           "Pon A1 bold=True e imprime font.bold.",
           "Font(bold=True).",
           "openpyxl.styles.",
           "from openpyxl import Workbook\nfrom openpyxl.styles import Font\n# TODO\n",
           """from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook()
ws = wb.active
ws["A1"] = "KPI"
ws["A1"].font = Font(bold=True)
print(ws["A1"].font.bold)""",
           ["estilo None"]),
        ex("S20-T2-A-E2", "S20-T2-A", "independent",
           "Aplica PatternFill solid fgColor 1F4E79 a A1 e imprime que fill no es None.",
           "PatternFill.",
           "print True.",
           "from openpyxl import Workbook\nfrom openpyxl.styles import PatternFill\n# TODO\n",
           """from openpyxl import Workbook
from openpyxl.styles import PatternFill
wb = Workbook()
ws = wb.active
ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
print(ws["A1"].fill.fgColor is not None)""",
           ["color theme vs rgb"]),
        ex("S20-T2-A-E3", "S20-T2-A", "transfer",
           "Escribe header_style(ws, cell) bold+fill; aplícalo a A1 y print bold.",
           "Función mutadora.",
           "Color corporativo.",
           "from openpyxl import Workbook\nfrom openpyxl.styles import Font, PatternFill\n# TODO\n",
           """from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def header_style(ws, coord):
    c = ws[coord]
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")

wb = Workbook()
ws = wb.active
ws["A1"] = "H"
header_style(ws, "A1")
print(ws["A1"].font.bold)""",
           ["celda inexistente"]),
        ex("S20-T2-B-E1", "S20-T2-B", "guided",
           "Escribe date(2024,1,15) en A1 e imprime isoformat.",
           "datetime.date.",
           "value.isoformat().",
           "from openpyxl import Workbook\nfrom datetime import date\n# TODO\n",
           """from openpyxl import Workbook
from datetime import date
wb = Workbook()
ws = wb.active
ws["A1"] = date(2024, 1, 15)
print(ws["A1"].value.isoformat())""",
           ["datetime vs date"]),
        ex("S20-T2-B-E2", "S20-T2-B", "independent",
           "merge B1:C1, valor en B1 'x'; imprime C1 value (None esperado).",
           "merge_cells.",
           "Leer no ancla.",
           "from openpyxl import Workbook\n# TODO\n",
           """from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.merge_cells("B1:C1")
ws["B1"] = "x"
print(ws["C1"].value)""",
           ["escribir en no ancla"]),
        ex("S20-T2-B-E3", "S20-T2-B", "transfer",
           "Cuenta cuántos merge ranges hay tras merge A1:B1 y C1:D1; print int.",
           "len(ws.merged_cells.ranges).",
           "Dos merges.",
           "from openpyxl import Workbook\n# TODO\n",
           """from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.merge_cells("A1:B1")
ws.merge_cells("C1:D1")
print(len(ws.merged_cells.ranges))""",
           ["unmerge"]),
        ex("S20-T3-A-E1", "S20-T3-A", "guided",
           "detalle sum 10+5, portada 15; print reconcile True con tol 0.01.",
           "abs(diff)<0.01.",
           "print bool.",
           "# TODO\n",
           """det = 10 + 5
portada = 15
print(abs(det - portada) < 0.01)""",
           ["floats 0.1+0.2"]),
        ex("S20-T3-A-E2", "S20-T3-A", "independent",
           "Pivot suma por región con pandas; print dict de Lima/Cusco para montos 10,5 y 7.",
           "groupby sum.",
           "to_dict.",
           "import pandas as pd\n# TODO\n",
           """import pandas as pd
df = pd.DataFrame({"region": ["Lima", "Lima", "Cusco"], "monto": [10.0, 5.0, 7.0]})
print(df.groupby("region")["monto"].sum().to_dict())""",
           ["NaN monto"]),
        ex("S20-T3-A-E3", "S20-T3-A", "transfer",
           "reconcile(det_sum, portada, tol=0.01) devuelve bool; prueba 22.0 vs 22.005 y 22.0 vs 23.",
           "abs <= tol o <.",
           "Dos prints.",
           "# TODO\n",
           """def reconcile(det_sum, portada, tol=0.01):
    return abs(det_sum - portada) < tol
print(reconcile(22.0, 22.005))
print(reconcile(22.0, 23.0))""",
           ["tol 0"]),
        ex("S20-T3-B-E1", "S20-T3-B", "guided",
           "headers got vs expected region/monto; print equality.",
           "list compare.",
           "True.",
           "expected=['region','monto']; got=['region','monto']\n# TODO\n",
           """expected = ["region", "monto"]
got = ["region", "monto"]
print(expected == got)""",
           ["orden distinto"]),
        ex("S20-T3-B-E2", "S20-T3-B", "independent",
           "Filtra regiones no permitidas de ['Lima','Piura'] con allowed Lima/Cusco; print bad list.",
           "list comp.",
           "print.",
           "# TODO\n",
           """allowed = {"Lima", "Cusco"}
regs = ["Lima", "Piura"]
print([r for r in regs if r not in allowed])""",
           ["case sensitivity"]),
        ex("S20-T3-B-E3", "S20-T3-B", "transfer",
           "validate_rows(rows, allowed) retorna lista de regiones inválidas; prueba una inválida.",
           "Comprensión.",
           "print resultado.",
           "# TODO\n",
           """def validate_rows(rows, allowed):
    return [r["region"] for r in rows if r["region"] not in allowed]
print(validate_rows([{"region": "Lima"}, {"region": "Ica"}], {"Lima", "Cusco"}))""",
           ["rows vacías"]),
        ex("S20-T4-A-E1", "S20-T4-A", "guided",
           "De statuses ok/corrupt/ok cuenta ok; print 2.",
           "sum genexp.",
           "print.",
           "st=['ok','corrupt','ok']\n# TODO\n",
           """st = ["ok", "corrupt", "ok"]
print(sum(x == "ok" for x in st))""",
           ["typos status"]),
        ex("S20-T4-A-E2", "S20-T4-A", "independent",
           "Clasifica nombre 'file.xlsx.lock' como locked si endswith .lock o contiene 'lock'; print status.",
           "Regla simple didáctica.",
           "print locked.",
           "name='report.lock'\n# TODO\n",
           """name = "report.lock"
print("locked" if name.endswith(".lock") or "lock" in name else "ok")""",
           ["falsos positivos"]),
        ex("S20-T4-A-E3", "S20-T4-A", "transfer",
           "batch_status(files_dict) devuelve conteos por estado; input {'a':'ok','b':'corrupt'}.",
           "Counter o dict.",
           "print dict.",
           "from collections import Counter\n# TODO\n",
           """from collections import Counter
files = {"a": "ok", "b": "corrupt"}
print(dict(Counter(files.values())))""",
           ["estado desconocido"]),
        ex("S20-T4-B-E1", "S20-T4-B", "guided",
           "Imprime manifest con backup path y idempotent True.",
           "dict.",
           "json opcional.",
           "# TODO\n",
           """print({"backup": "out/prev.bak", "idempotent": True})""",
           ["path absoluto con secretos"]),
        ex("S20-T4-B-E2", "S20-T4-B", "independent",
           "Muestra que build ordenado es idempotente: same hash para rows en distinto orden.",
           "sorted + sha1.",
           "print True.",
           "import hashlib\n# TODO\n",
           """import hashlib

def dig(rows):
    s = "\\n".join(f"{a},{b}" for a, b in sorted(rows))
    return hashlib.sha1(s.encode()).hexdigest()
print(dig([("Lima", 1), ("Cusco", 2)]) == dig([("Cusco", 2), ("Lima", 1)]))""",
           ["floats formatting"]),
        ex("S20-T4-B-E3", "S20-T4-B", "transfer",
           "structural_ok(sheetnames, need) True si need subset de sheetnames; prueba need Entrada/Salida.",
           "set issuperset.",
           "print.",
           "# TODO\n",
           """def structural_ok(sheetnames, need):
    return set(sheetnames) >= set(need)
print(structural_ok(["Entrada", "Salida", "Log"], ["Entrada", "Salida"]))""",
           ["case"]),
    ]

    youdo = {
        "title": "Excel factory CP-N2-B",
        "context": (
            "Implementa un adaptador que lea formatos sintéticos del VP, genere workbook de resultados sin dañar "
            "la plantilla master y deje **manifest** de cambios/conciliación (**CP-N2-B excel factory**)."
        ),
        "objectives": [
            "Leer/escribir sheets con openpyxl",
            "Calcular en Python valores de salida",
            "Conciliar totales con tolerancia",
            "Batch con corrupt/lock handling",
            "Backup + idempotencia + tests estructurales",
        ],
        "requirements": [
            "No modificar plantilla master in-place sin backup",
            "Manifest JSON con estados y reconcile_ok",
            "Datos sintéticos only",
            "Headers validados",
            "es-PE en etiquetas de hojas de presentación",
        ],
        "starterCode": """from openpyxl import Workbook
import json

# TODO: plantilla → salida + manifest
wb = Workbook()
ws = wb.active
ws.title = "Entrada"
ws.append(["region", "monto"])
ws.append(["Lima", 10.0])
print(wb.sheetnames)
""",
        "portfolioNote": "Workbook + manifest del factory; se enlaza al dashboard (S19) y al paquete de reportes (S21).",
        "rubric": RUBRIC_STD,
    }

    selfcheck = [
        {
            "question": "openpyxl sin Excel instalado evalúa fórmulas automáticamente:",
            "options": ["Siempre sí", "No; suele devolver la fórmula o cache si existe", "Solo en Linux", "Solo named ranges"],
            "correctIndex": 1,
            "explanation": "No hay motor Excel en openpyxl por defecto; calcula en Python o usa data_only con cache previo.",
        },
        {
            "question": "Al escribir en celdas combinadas debes:",
            "options": ["Escribir en cualquier celda del merge", "Escribir en la celda ancla (top-left)", "Desmerge siempre", "Usar solo CSV"],
            "correctIndex": 1,
            "explanation": "El valor vive en la celda ancla del rango combinado.",
        },
        {
            "question": "Un manifest del excel factory debe permitir auditar:",
            "options": ["Solo el color de fuente", "Estados de batch, conciliación y backups", "La contraseña del VP", "Embeddings"],
            "correctIndex": 1,
            "explanation": "Auditoría operativa del lote y de la conciliación.",
        },
        {
            "question": "Idempotencia significa:",
            "options": ["Correr dos veces cambia totales al azar", "Misma entrada → mismo resultado lógico", "Borrar la plantilla", "Ignorar headers"],
            "correctIndex": 1,
            "explanation": "Re-ejecutar no debe corromper ni duplicar efectos no controlados.",
        },
    ]

    resources = {
        "docs": [
            {"label": "openpyxl docs", "url": "https://openpyxl.readthedocs.io/", "note": "API de workbooks y estilos"},
            {"label": "Office Open XML overview", "url": "https://learn.microsoft.com/en-us/office/open-xml/open-xml-sdk", "note": "Contexto de formato xlsx"},
        ],
        "books": [
            {"label": "Automate the Boring Stuff (Excel chapters)", "note": "Automatización práctica"},
            {"label": "Python for Excel (Zumstein)", "note": "Patrones profesionales con openpyxl/xlwings"},
        ],
        "courses": [
            {"label": "openpyxl tutorial", "url": "https://openpyxl.readthedocs.io/en/stable/tutorial.html", "note": "Inicio rápido"},
        ],
    }

    i_do = "Te demuestro el excel factory: sheets, fórmulas vs Python, estilos, merges, conciliación, batch y manifest."
    we_do = "24 ejercicios con openpyxl y pandas para un adaptador robusto de plantillas sintéticas."
    ts, log = build_section(
        "section20", meta, theories, demos, exercises, i_do, we_do, youdo, selfcheck, resources
    )
    meta_updates = {
        "id": "rag",
        "index": 20,
        "title": meta["title"],
        "shortTitle": meta["shortTitle"],
        "tagline": meta["tagline"],
        "estimatedHours": 12,
        "learningOutcomes": 8,
        "youDo": youdo["title"],
        "icon": "MessageSquare",
        "legacy_note": "RAG demoted; target is robust Excel automation for CP-N2-B excel factory",
        "capstone": "CP-N2-B (excel factory)",
    }
    return ts, log, slugs, meta_updates, youdo["title"]



# ---------------------------------------------------------------------------
# S21 — Documentos, plantillas y reportes trazables (platform id: fastapi)
# ---------------------------------------------------------------------------


def build_s21() -> tuple[str, dict[str, str], list[tuple[str, str]], dict, str]:
    slugs = [
        ("S21-T1-A", "jinja-data-presentation"),
        ("S21-T1-B", "conditions-tables-safe-format"),
        ("S21-T2-A", "styles-sections-docx-pdf"),
        ("S21-T2-B", "pdf-digital-vs-image-ocr"),
        ("S21-T3-A", "exec-summary-method-findings"),
        ("S21-T3-B", "charts-tables-sources-limits"),
        ("S21-T4-A", "copy-a11y-consistency"),
        ("S21-T4-B", "visual-render-provenance-approval"),
    ]
    meta = {
        "id": "fastapi",
        "index": 21,
        "title": "Documentos, plantillas y reportes trazables",
        "shortTitle": "Reportes trazables",
        "tagline": "Accessible Insights Dashboard & Reporting Factory genera dashboard, DOCX/PDF y workbook desde una corrida, con números reconciliados y revisión visual",
        "estimatedHours": 14,
        "level": "Competente",
        "phase": 1,
        "icon": "Server",
        "accentColor": "bg-gradient-to-br from-blue-500 to-indigo-600",
        "jobRelevance": (
            "Cerrar **CP-N2-B** exige un **Reporting Factory** que una dashboard, Excel y documentos con números "
            "reconciliados y provenance. Esta sección (id `fastapi` conservado) retematiza a V3 **Documentos, "
            "plantillas y reportes trazables** con Jinja, estructura ejecutiva y cola de aprobación."
        ),
        "learningOutcomes": [
            "Separar datos y presentación con Jinja",
            "Renderizar condiciones/tablas con formato seguro",
            "Generar estructuras tipo DOCX/PDF con secciones",
            "Distinguir PDF digital de imagen/OCR",
            "Estructurar informe ejecutivo con método y hallazgos",
            "Embeber gráficos/tablas con fuentes y limitaciones",
            "Revisar redacción, a11y y consistencia numérica",
            "Documentar provenance y flujo de aprobación",
        ],
    }

    theories = [
        Theory(
            heading="De “FastAPI backend” a Reporting Factory (mapa y cierre CP-N2-B)",
            paragraphs=[
                "En V3, **S21 no es el path principal de FastAPI/OpenAPI**. Aquí **cierras CP-N2-B**: plantillas Jinja, reportes con secciones, consistencia numérica con dashboard/Excel y **provenance + aprobación**.",
                "Una sola corrida produce artefactos alineados (mismos n y métricas que S18–S20).",
                "Orden: **T1 Plantillas** → **T2 Documentos** → **T3 Narrativa** → **T4 Calidad y gobernanza**.",
            ],
            callout=(
                "info",
                "Contenido reubicado conceptualmente",
                "Material legado de FastAPI de este archivo **no es el camino V3 en S21**. Target: reportes trazables y cierre CP-N2-B.",
            ),
        ),
        Theory(
            heading="Jinja y separación datos/presentación",
            sub="S21-T1-A",
            paragraphs=[
                "Jinja permite **plantillas** con `{{ variables }}` y `{% bloques %}`. Los datos se preparan en Python; la plantilla solo presenta.",
                "Nunca pongas lógica de negocio pesada en la plantilla: calcula métricas antes del render.",
                "Autoescape en HTML; en texto plano define políticas de caracteres.",
            ],
            code="""from jinja2 import Template

tmpl = Template("Región {{ region }}: mediana {{ median }} PEN (n={{ n }})")
print(tmpl.render(region="Lima", median=28.0, n=40))""",
            code_title="jinja_basic.py",
            callout=(
                "tip",
                "Context dict único",
                "Pasa un context versionado (run_id, metricas, limites) a todas las plantillas del factory.",
            ),
        ),
        Theory(
            heading="condiciones, tablas y formato seguro",
            sub="S21-T1-B",
            paragraphs=[
                "`{% if %}` y `{% for %}` construyen tablas. Formatea números con filtros o preformateo en Python (`f\"{x:.2f}\"`) para consistencia.",
                "Evita inyección: no marques strings de usuario como safe en HTML sin sanitizar.",
                "Celdas vacías: muestra “—” y documenta missing, no inventes ceros.",
            ],
            code="""from jinja2 import Template

tmpl = Template(
    \"\"\"{% for r in rows %}- {{ r.region }}: {{ '%.2f'|format(r.median) }} PEN
{% endfor %}\"\"\"
)
rows = [{"region": "Lima", "median": 28.0}, {"region": "Cusco", "median": 22.5}]
print(tmpl.render(rows=rows))""",
            code_title="jinja_table.py",
            callout=(
                "warning",
                "Cero vs missing",
                "Imprimir 0.00 cuando no hay datos es un error de reporting grave.",
            ),
        ),
        Theory(
            heading="estilos y secciones (DOCX/PDF conceptual)",
            sub="S21-T2-A",
            paragraphs=[
                "Un informe trazable tiene secciones fijas: portada, resumen, método, hallazgos, anexos. Los estilos (Heading 1/2) habilitan TOC y a11y.",
                "En este curso modelamos el documento como estructura de bloques renderizada a texto/Markdown/HTML; la migración a python-docx es directa.",
                "Misma jerarquía en PDF digital facilita extracción posterior.",
            ],
            code="""doc = {
    "title": "Informe de insights sintético",
    "sections": [
        {"h": 1, "name": "Resumen ejecutivo"},
        {"h": 2, "name": "Método"},
        {"h": 2, "name": "Hallazgos"},
        {"h": 1, "name": "Anexos"},
    ],
}
print([ (s["h"], s["name"]) for s in doc["sections"] ])
print("h1_count", sum(1 for s in doc["sections"] if s["h"] == 1))""",
            code_title="sections.py",
            callout=(
                "tip",
                "Outline primero",
                "Congela el outline antes de redactar párrafos largos.",
            ),
        ),
        Theory(
            heading="PDF digital vs imagen/OCR",
            sub="S21-T2-B",
            paragraphs=[
                "Un **PDF digital** contiene texto seleccionable; un **PDF escaneado** es imagen y requiere OCR con errores típicos.",
                "Clasifica antes de extraer: si no hay texto o el ratio de caracteres es bajo, marca `needs_ocr`.",
                "En el factory, prefiere generar PDF digital desde HTML/LaTeX/reportlab en lugar de escanear.",
            ],
            code="""def classify_pdf(text_layer_chars: int, pages: int) -> str:
    density = text_layer_chars / max(pages, 1)
    if density < 20:
        return "image_or_scan_needs_ocr"
    return "digital_text"
print(classify_pdf(5000, 3))
print(classify_pdf(10, 3))""",
            code_title="pdf_class.py",
            callout=(
                "info",
                "OCR no es verdad absoluta",
                "Todo número OCR-eado debe reconciliarse con la fuente tabular cuando exista.",
            ),
        ),
        Theory(
            heading="resumen ejecutivo, método y hallazgos",
            sub="S21-T3-A",
            paragraphs=[
                "El **resumen ejecutivo** responde la pregunta de negocio en ≤5 viñetas con n e incertidumbre. **Método** describe datos, filtros y métricas. **Hallazgos** citan tablas/figuras.",
                "No mezcles método con opinión. Cada hallazgo mapea a un id de evidencia (p. ej. H1 → Tabla 2).",
                "Alineado a S18: hallazgo ≠ decisión.",
            ],
            code="""report = {
    "resumen": ["Ticket mediano Lima 28 PEN en muestra web (n=40)"],
    "metodo": {"fuente": "sintetico", "filtros": ["canal=web"], "corte": "2024-06-30"},
    "hallazgos": [{"id": "H1", "texto": "Lima > Cusco en mediana", "evidencia": "Tabla2"}],
}
print(report["resumen"][0])
print(report["hallazgos"][0]["evidencia"])""",
            code_title="exec_struct.py",
            callout=(
                "success",
                "Trazabilidad H→evidencia",
                "Sin id de evidencia, el hallazgo no entra al paquete de aprobación.",
            ),
        ),
        Theory(
            heading="gráficos, tablas, fuentes y limitaciones",
            sub="S21-T3-B",
            paragraphs=[
                "Inserta figuras del dashboard (S19) y tablas del Excel (S20) con **caption** idéntico en fuente/corte/n.",
                "Lista de limitaciones al final de hallazgos, no escondida en anexo solo.",
                "Reconciliación: checksum de métricas clave entre artefactos.",
            ],
            code="""metrics = {"median_Lima": 28.0, "n_Lima": 40}
caption = "Fig.1 Ticket mediano | Fuente: sintético | n_Lima=40"
assert "40" in caption
bundle = {"fig": "fig1.png", "table": "tabla2", "metrics": metrics, "limits": ["solo web"]}
print(bundle["metrics"])
print(bundle["limits"])""",
            code_title="embed_limits.py",
            callout=(
                "warning",
                "Números divergentes",
                "Si el DOCX dice 28 y el Excel 27.5, el factory falla el gate de cierre.",
            ),
        ),
        Theory(
            heading="redacción, accesibilidad y consistencia",
            sub="S21-T4-A",
            paragraphs=[
                "Redacción clara en español profesional (es-PE): evita anglicismos innecesarios en el cuerpo ejecutivo.",
                "A11y: headings reales, texto alt de figuras, tablas con encabezados, contraste en HTML.",
                "Consistencia: misma precisión decimal (p. ej. 1 decimal PEN) en todo el paquete.",
            ],
            code="""vals = [28.04, 28.0, 28]
precision = 1
norm = [round(float(v), precision) for v in vals]
print(norm)
print("consistente", len(set(norm)) == 1)""",
            code_title="consistency.py",
            callout=(
                "tip",
                "Una función format_metric",
                "Centraliza redondeo y unidades para no divergir entre Jinja y Excel.",
            ),
        ),
        Theory(
            heading="render visual, provenance y aprobación",
            sub="S21-T4-B",
            paragraphs=[
                "Registra **provenance**: run_id, git/sha de datos, versiones de script, hashes de artefactos.",
                "Cola de **aprobación**: borrador → revisión visual → aprobado/rechazado con comentarios.",
                "El cierre de CP-N2-B exige provenance completo + checklist visual (dashboard, xlsx, doc).",
            ],
            code="""import hashlib, json
from datetime import datetime, timezone

artifacts = {"dashboard": "ok", "xlsx": "ok", "doc": "ok"}
prov = {
    "run_id": "cpn2b-20240630-01",
    "ts": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    "data_sha1_8": hashlib.sha1(b"synthetic").hexdigest()[:8],
    "artifacts": artifacts,
    "approval": {"status": "pending_review", "reviewer": None},
}
print(json.dumps(prov, ensure_ascii=False))""",
            code_title="provenance.py",
            callout=(
                "success",
                "Gate de cierre",
                "Sin provenance y sin revisión visual registrada, CP-N2-B no se considera cerrado en contenido (el ledger lo confirma otra lane).",
            ),
        ),
    ]

    demos = [
        Demo(
            "S21-T1-A-DEMO", "S21-T1-A",
            "Separar datos de plantilla Jinja con context único",
            """from jinja2 import Template

context = {
    "run_id": "r1",
    "region": "Lima",
    "median": 28.0,
    "n": 40,
    "limit": "solo web",
}
tmpl = Template(
    "Run {{ run_id }} — {{ region }}: {{ median }} PEN (n={{ n }}). Límite: {{ limit }}."
)
print(tmpl.render(**context))""",
            "Un context versionado alimenta todos los renders del factory.",
            title="demo_jinja.py",
        ),
        Demo(
            "S21-T1-B-DEMO", "S21-T1-B",
            "Render condicional seguro de tabla con missing como em-dash",
            """from jinja2 import Template

tmpl = Template(
    \"\"\"{% for r in rows %}{{ r.region }}: {{ r.median if r.median is not none else '—' }}
{% endfor %}\"\"\"
)
rows = [
    {"region": "Lima", "median": 28.0},
    {"region": "Cusco", "median": None},
]
print(tmpl.render(rows=rows))""",
            "Missing explícito evita ceros inventados en el informe.",
            title="demo_cond_table.py",
        ),
        Demo(
            "S21-T2-A-DEMO", "S21-T2-A",
            "Generar outline de secciones estilo DOCX/PDF",
            """def render_outline(sections):
    lines = []
    for s in sections:
        pad = "  " * (s["h"] - 1)
        lines.append(f"{pad}H{s['h']} {s['name']}")
    return "\\n".join(lines)

sections = [
    {"h": 1, "name": "Resumen ejecutivo"},
    {"h": 1, "name": "Método"},
    {"h": 2, "name": "Datos y filtros"},
    {"h": 1, "name": "Hallazgos"},
]
print(render_outline(sections))
print("sections", len(sections))""",
            "El outline fija la estructura del paquete documental.",
            title="demo_sections.py",
        ),
        Demo(
            "S21-T2-B-DEMO", "S21-T2-B",
            "Clasificar PDF digital vs posible escaneo por densidad de texto",
            """def classify(chars, pages):
    dens = chars / max(pages, 1)
    return ("digital", dens) if dens >= 20 else ("needs_ocr", dens)

print(classify(9000, 4))
print(classify(15, 4))""",
            "Clasificar antes de extraer evita basar KPIs en OCR ruidoso.",
            title="demo_pdf_class.py",
        ),
        Demo(
            "S21-T3-A-DEMO", "S21-T3-A",
            "Estructurar informe en resumen, método y hallazgos con ids",
            """informe = {
    "resumen": [
        "En muestra web sintética, Lima tiene ticket mediano 28 PEN (n=40).",
    ],
    "metodo": {
        "fuente": "dataset sintético S18",
        "filtros": ["monto>0", "canal=web"],
    },
    "hallazgos": [
        {"id": "H1", "claim": "Lima > Cusco en mediana", "evidencia": "Tabla2", "decision": None},
    ],
}
print(informe["resumen"][0])
print(informe["hallazgos"][0]["id"], informe["hallazgos"][0]["evidencia"])
print("decision_none", informe["hallazgos"][0]["decision"] is None)""",
            "Ids de hallazgo habilitan revisión y aprobación selectiva.",
            title="demo_exec.py",
        ),
        Demo(
            "S21-T3-B-DEMO", "S21-T3-B",
            "Embeber métricas con fuentes/límites y check de paridad",
            """dash = {"median_Lima": 28.0}
xlsx = {"median_Lima": 28.0}
doc = {"median_Lima": 28.0}
limits = ["cobertura web", "n Cusco bajo"]
parity = dash == xlsx == doc
bundle = {"parity": parity, "limits": limits, "fuente": "sintetico"}
print(bundle)""",
            "Paridad entre dashboard, Excel y documento es el corazón del cierre.",
            title="demo_parity.py",
        ),
        Demo(
            "S21-T4-A-DEMO", "S21-T4-A",
            "Normalizar decimales y validar presencia de headings/alt",
            """def fmt_pen(x):
    return f"{round(float(x), 1)} PEN"

checks = {
    "decimals": [fmt_pen(28.04), fmt_pen(28.0)],
    "has_h1": True,
    "alts": ["Barras mediana por región, n por barra en tooltip"],
}
print(checks["decimals"])
print("decimal_ok", len(set(checks["decimals"])) == 1)
print("a11y_min", checks["has_h1"] and all(len(a) > 10 for a in checks["alts"]))""",
            "Consistencia tipográfica y a11y mínima antes de mandar a revisión.",
            title="demo_a11y_copy.py",
        ),
        Demo(
            "S21-T4-B-DEMO", "S21-T4-B",
            "Registrar provenance y estado de cola de aprobación",
            """import hashlib, json

prov = {
    "run_id": "cpn2b-close-01",
    "data_sha1_8": hashlib.sha1(b"rows-synthetic").hexdigest()[:8],
    "artifacts": ["dashboard.html", "results.xlsx", "informe.md"],
    "visual_checklist": {"dashboard": True, "xlsx": True, "doc": True},
    "approval": {"status": "pending_review"},
}
print(json.dumps(prov, ensure_ascii=False))
print("ready_for_review", all(prov["visual_checklist"].values()))""",
            "Provenance + checklist visual cierran el Reporting Factory CP-N2-B.",
            title="demo_prov.py",
        ),
    ]

    exercises = [
        ex("S21-T1-A-E1", "S21-T1-A", "guided",
           "Renderiza Template('Hola {{ nombre }}') con nombre='Ana' e imprime.",
           "jinja2.Template.render.",
           "print.",
           "from jinja2 import Template\n# TODO\n",
           """from jinja2 import Template
print(Template("Hola {{ nombre }}").render(nombre="Ana"))""",
           ["nombre missing → vacío"]),
        ex("S21-T1-A-E2", "S21-T1-A", "independent",
           "Template con median y n; render Lima-like 28 y 40.",
           "Dos variables.",
           "Incluye PEN.",
           "from jinja2 import Template\n# TODO\n",
           """from jinja2 import Template
print(Template("{{ m }} PEN (n={{ n }})").render(m=28, n=40))""",
           ["tipos str vs int"]),
        ex("S21-T1-A-E3", "S21-T1-A", "transfer",
           "Función render_kpi(ctx) usa template fijo region/median/n; prueba con dict.",
           "Template dentro o global.",
           "print resultado.",
           "from jinja2 import Template\n# TODO\n",
           """from jinja2 import Template

def render_kpi(ctx):
    return Template("{{ region }}: {{ median }} PEN (n={{ n }})").render(**ctx)
print(render_kpi({"region": "Cusco", "median": 22.5, "n": 32}))""",
           ["key error"]),
        ex("S21-T1-B-E1", "S21-T1-B", "guided",
           "Si median is None imprime '—'; else el número. Prueba None.",
           "condicional python o jinja.",
           "print.",
           "median = None\n# TODO\n",
           """median = None
print("—" if median is None else median)""",
           ["NaN float"]),
        ex("S21-T1-B-E2", "S21-T1-B", "independent",
           "Formatea 28.456 a 2 decimales con format en jinja o python; print.",
           "'%.2f' % x o f-string.",
           "print 28.46.",
           "x = 28.456\n# TODO\n",
           """x = 28.456
print(f"{x:.2f}")""",
           ["locale comma"]),
        ex("S21-T1-B-E3", "S21-T1-B", "transfer",
           "Render lista de rows en líneas 'region:value' con jinja for.",
           "Template for.",
           "Dos regiones.",
           "from jinja2 import Template\n# TODO\n",
           """from jinja2 import Template
tmpl = Template("{% for r in rows %}{{ r.region }}:{{ r.v }}\\n{% endfor %}")
print(tmpl.render(rows=[{"region": "Lima", "v": 1}, {"region": "Cusco", "v": 2}]), end="")""",
           ["rows vacías"]),
        ex("S21-T2-A-E1", "S21-T2-A", "guided",
           "Define sections con Resumen h1 y Método h1; imprime nombres.",
           "lista de dicts.",
           "print list comprehension.",
           "# TODO\n",
           """sections = [{"h": 1, "name": "Resumen"}, {"h": 1, "name": "Método"}]
print([s["name"] for s in sections])""",
           ["h inválido"]),
        ex("S21-T2-A-E2", "S21-T2-A", "independent",
           "Cuenta cuántos H1 hay en sections con h 1 y 2 mixtos.",
           "sum.",
           "print int.",
           "sections=[{'h':1},{'h':2},{'h':1}]\n# TODO\n",
           """sections = [{"h": 1}, {"h": 2}, {"h": 1}]
print(sum(1 for s in sections if s["h"] == 1))""",
           ["h string"]),
        ex("S21-T2-A-E3", "S21-T2-A", "transfer",
           "render_outline devuelve líneas indentadas; prueba un H1 y un H2.",
           "pad por h-1.",
           "print.",
           "# TODO\n",
           """def render_outline(sections):
    return "\\n".join("  " * (s["h"] - 1) + s["name"] for s in sections)
print(render_outline([{"h": 1, "name": "A"}, {"h": 2, "name": "B"}]))""",
           ["h=0"]),
        ex("S21-T2-B-E1", "S21-T2-B", "guided",
           "Si chars/pages < 20 print needs_ocr else digital; caso 10 chars 2 pages.",
           "densidad.",
           "print.",
           "chars, pages = 10, 2\n# TODO\n",
           """chars, pages = 10, 2
dens = chars / pages
print("needs_ocr" if dens < 20 else "digital")""",
           ["pages 0"]),
        ex("S21-T2-B-E2", "S21-T2-B", "independent",
           "classify(chars,pages) retorna digital o needs_ocr; prueba (100,1).",
           "Función.",
           "print.",
           "# TODO\n",
           """def classify(chars, pages):
    return "digital" if chars / max(pages, 1) >= 20 else "needs_ocr"
print(classify(100, 1))""",
           ["umbral distinto"]),
        ex("S21-T2-B-E3", "S21-T2-B", "transfer",
           "Dado texto extraído '', marca needs_ocr True; imprime dict flags.",
           "bool(text).",
           "print.",
           "text = ''\n# TODO\n",
           """text = ""
print({"needs_ocr": not bool(text.strip()), "n_chars": len(text)})""",
           ["solo espacios"]),
        ex("S21-T3-A-E1", "S21-T3-A", "guided",
           "Crea hallazgo id H1 con evidencia Tabla1 e imprime id.",
           "dict.",
           "print.",
           "# TODO\n",
           """h = {"id": "H1", "evidencia": "Tabla1"}
print(h["id"])""",
           ["id duplicado"]),
        ex("S21-T3-A-E2", "S21-T3-A", "independent",
           "Resumen debe incluir 'n='; valida e imprime True/False para un string dado.",
           "in.",
           "caso con n=40.",
           "s = 'mediana 28 PEN n=40'\n# TODO\n",
           """s = "mediana 28 PEN n=40"
print("n=" in s)""",
           ["N mayúscula"]),
        ex("S21-T3-A-E3", "S21-T3-A", "transfer",
           "pack_report(resumen, metodo, hallazgos) devuelve dict con 3 claves; print keys sorted.",
           "dict function.",
           "print.",
           "# TODO\n",
           """def pack_report(resumen, metodo, hallazgos):
    return {"resumen": resumen, "metodo": metodo, "hallazgos": hallazgos}
print(sorted(pack_report(["a"], {}, []).keys()))""",
           ["tipos"]),
        ex("S21-T3-B-E1", "S21-T3-B", "guided",
           "parity entre dash y doc dicts median_Lima 28; print True.",
           "comparar.",
           "print.",
           "# TODO\n",
           """dash = {"median_Lima": 28.0}
doc = {"median_Lima": 28.0}
print(dash == doc)""",
           ["float vs int"]),
        ex("S21-T3-B-E2", "S21-T3-B", "independent",
           "Caption debe contener 'Fuente'; valida.",
           "in.",
           "print bool.",
           "cap = 'Fig1 | Fuente: sintetico | n=10'\n# TODO\n",
           """cap = "Fig1 | Fuente: sintetico | n=10"
print("Fuente" in cap)""",
           ["fuente minúscula"]),
        ex("S21-T3-B-E3", "S21-T3-B", "transfer",
           "check_parity(a,b,c) True si los tres iguales; prueba fallo y éxito.",
           "a==b==c.",
           "Dos prints.",
           "# TODO\n",
           """def check_parity(a, b, c):
    return a == b == c
print(check_parity({"x": 1}, {"x": 1}, {"x": 1}))
print(check_parity({"x": 1}, {"x": 1}, {"x": 2}))""",
           ["keys extra"]),
        ex("S21-T4-A-E1", "S21-T4-A", "guided",
           "Normaliza [28.04, 28.0] a 1 decimal; print lista.",
           "round.",
           "print.",
           "vals=[28.04,28.0]\n# TODO\n",
           """vals = [28.04, 28.0]
print([round(v, 1) for v in vals])""",
           ["banker's rounding"]),
        ex("S21-T4-A-E2", "S21-T4-A", "independent",
           "fmt_pen(28.04) -> '28.0 PEN'; implementa e imprime.",
           "f-string round 1.",
           "print.",
           "# TODO\n",
           """def fmt_pen(x):
    return f"{round(float(x), 1)} PEN"
print(fmt_pen(28.04))""",
           ["None"]),
        ex("S21-T4-A-E3", "S21-T4-A", "transfer",
           "a11y_min(has_h1, alts) True si has_h1 y todo alt len>10; prueba.",
           "all().",
           "print.",
           "# TODO\n",
           """def a11y_min(has_h1, alts):
    return has_h1 and all(len(a) > 10 for a in alts)
print(a11y_min(True, ["descripcion larga de figura"]))
print(a11y_min(True, ["corto"]))""",
           ["alts vacía"]),
        ex("S21-T4-B-E1", "S21-T4-B", "guided",
           "Imprime approval status 'pending_review'.",
           "dict.",
           "print value.",
           "# TODO\n",
           """approval = {"status": "pending_review"}
print(approval["status"])""",
           ["typo status"]),
        ex("S21-T4-B-E2", "S21-T4-B", "independent",
           "data_sha1_8 de b'synthetic' primeros 8 hex; print.",
           "hashlib.sha1.",
           "print slice.",
           "import hashlib\n# TODO\n",
           """import hashlib
print(hashlib.sha1(b"synthetic").hexdigest()[:8])""",
           ["encoding"]),
        ex("S21-T4-B-E3", "S21-T4-B", "transfer",
           "ready(checklist) True si todos los valores True; prueba dashboard/xlsx/doc.",
           "all(dict.values()).",
           "print.",
           "# TODO\n",
           """def ready(checklist):
    return all(checklist.values())
print(ready({"dashboard": True, "xlsx": True, "doc": True}))
print(ready({"dashboard": True, "xlsx": False, "doc": True}))""",
           ["keys faltantes"]),
    ]

    youdo = {
        "title": "Reporting Factory — cierre CP-N2-B",
        "context": (
            "Integra EDA (S18), dashboard (S19) y Excel (S20) en un **Accessible Insights Dashboard & Reporting "
            "Factory**: una corrida genera documentos/plantillas con números reconciliados, alt text, provenance y "
            "cola de aprobación. Cierre de **CP-N2-B**."
        ),
        "objectives": [
            "Plantillas Jinja con context único",
            "Outline de informe ejecutivo",
            "Paridad de métricas entre artefactos",
            "A11y y formato consistente",
            "Provenance + pending_review",
        ],
        "requirements": [
            "Sin PII real ni secretos",
            "Hallazgos con id y evidencia",
            "Missing ≠ 0",
            "Checksum/paridad documentada",
            "es-PE en narrativa",
        ],
        "starterCode": """from jinja2 import Template
import json, hashlib

context = {
    "run_id": "cpn2b-01",
    "median_Lima": 28.0,
    "n_Lima": 40,
    "limits": ["solo web"],
}
# TODO: render informe + provenance
print(Template("Run {{ run_id }} Lima={{ median_Lima }}").render(**context))
""",
        "portfolioNote": "Paquete final CP-N2-B: dashboard + xlsx + informe con provenance; listo para revisión de gate (otra lane marca passed).",
        "rubric": RUBRIC_STD,
    }

    selfcheck = [
        {
            "question": "¿Por qué separar datos y plantilla Jinja?",
            "options": ["Para mezclar SQL en el HTML", "Para reutilizar presentación y auditar métricas en Python", "Para evitar control de versiones", "Para ocultar n"],
            "correctIndex": 1,
            "explanation": "La lógica y métricas viven en Python; la plantilla presenta.",
        },
        {
            "question": "Un PDF con casi sin caracteres en capa de texto suele requerir:",
            "options": ["Ignorar el archivo", "OCR / tratamiento de imagen", "Solo openpyxl", "Borrar limitaciones"],
            "correctIndex": 1,
            "explanation": "Baja densidad de texto sugiere escaneo/imagen.",
        },
        {
            "question": "Paridad en el Reporting Factory significa:",
            "options": ["Mismos colores", "Mismas métricas clave en dashboard, Excel y documento", "Mismo número de páginas", "Mismo reviewer"],
            "correctIndex": 1,
            "explanation": "Números reconciliados entre artefactos.",
        },
        {
            "question": "El cierre de contenido de CP-N2-B incluye:",
            "options": ["Solo un print", "Provenance, checklist visual y hallazgos trazables", "Subir secretos al repo", "Marcar section_passed desde esta lane"],
            "correctIndex": 1,
            "explanation": "Esta lane no marca passed en ledger; sí entrega artefactos con provenance.",
        },
    ]

    resources = {
        "docs": [
            {"label": "Jinja2 documentation", "url": "https://jinja.palletsprojects.com/", "note": "Templates y sandbox"},
            {"label": "WCAG brief", "url": "https://www.w3.org/WAI/standards-guidelines/wcag/", "note": "Accesibilidad de contenidos"},
        ],
        "books": [
            {"label": "Docs for Developers", "note": "Estructura y claridad de documentos técnicos"},
            {"label": "The Data Warehouse Toolkit (kimball) — select chapters", "note": "Narrativa de métricas y linaje (conceptual)"},
        ],
        "courses": [
            {"label": "Jinja primer", "url": "https://jinja.palletsprojects.com/en/stable/templates/", "note": "Sintaxis de plantillas"},
        ],
    }

    i_do = "Te muestro el Reporting Factory: Jinja, secciones, paridad de números, a11y y provenance para cerrar CP-N2-B."
    we_do = "24 ejercicios de plantillas, clasificación PDF, estructura ejecutiva y gobernanza de aprobación."
    ts, log = build_section(
        "section21", meta, theories, demos, exercises, i_do, we_do, youdo, selfcheck, resources
    )
    meta_updates = {
        "id": "fastapi",
        "index": 21,
        "title": meta["title"],
        "shortTitle": meta["shortTitle"],
        "tagline": meta["tagline"],
        "estimatedHours": 14,
        "learningOutcomes": 8,
        "youDo": youdo["title"],
        "icon": "Server",
        "legacy_note": "FastAPI demoted; target is traceable reports and CP-N2-B close",
        "capstone": "CP-N2-B (cierre)",
    }
    return ts, log, slugs, meta_updates, youdo["title"]


def main() -> None:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    builders = [
        ("S18", "data-engineering", "s18-data-engineering.ts", "section18", build_s18, "CP-N2-B (inicio)"),
        ("S19", "databases-orm", "s19-databases-orm.ts", "section19", build_s19, "CP-N2-B (dashboard)"),
        ("S20", "rag", "s20-rag.ts", "section20", build_s20, "CP-N2-B (excel factory)"),
        ("S21", "fastapi", "s21-fastapi.ts", "section21", build_s21, "CP-N2-B (cierre)"),
    ]
    results = {}
    for section, sid, fname, _export, builder, gate in builders:
        print(f"Building {section}…")
        ts, log, slugs, meta_u, youdo_t = builder()
        path = SECTIONS / fname
        path.write_text(ts, encoding="utf-8")
        print("Wrote", path, "verified", len(log))
        prog = progress_payload(section, sid, fname, meta_u, slugs, log, youdo_t)
        prog_path = STATE / f"{section.lower()}_phase4_progress.json"
        prog_path.write_text(json.dumps(prog, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        results[section] = {
            "section_id": sid,
            "title_v3": meta_u["title"],
            "subtopics_done": [s[0] for s in slugs],
            "subtopics_remaining": [],
            "exercises_done": 24,
            "demos_done": 8,
            "gate": gate,
            "log": log,
            "fname": fname,
        }

    lane = {
        "lane_id": LANE_ID,
        "parent_lane": None,
        "sections": ["S18", "S19", "S20", "S21"],
        "mode": "PARALLEL_PRODUCTION",
        "role": "Author Agent PHASE 4",
        "status": "PHASE_4_COMPLETE",
        "section_passed": False,
        "updated_at": now,
        "S18": {k: v for k, v in results["S18"].items() if k != "log"},
        "S19": {k: v for k, v in results["S19"].items() if k != "log"},
        "S20": {k: v for k, v in results["S20"].items() if k != "log"},
        "S21": {k: v for k, v in results["S21"].items() if k != "log"},
        "exercises_done": 96,
        "exercises_target": 96,
        "demos_done": 32,
        "demos_target": 32,
        "files_changed": [
            "src/lib/course/sections/s18-data-engineering.ts",
            "src/lib/course/sections/s19-databases-orm.ts",
            "src/lib/course/sections/s20-rag.ts",
            "src/lib/course/sections/s21-fastapi.ts",
            "course-state/s18_phase4_progress.json",
            "course-state/s19_phase4_progress.json",
            "course-state/s20_phase4_progress.json",
            "course-state/s21_phase4_progress.json",
            f"course-state/lanes/{LANE_ID}.status.json",
            "scripts/_gen_s18_s21_p4.py",
        ],
        "execution_summary": (
            "Retargeted S18→EDA e incertidumbre (CP-N2-B inicio), S19→viz accesible (dashboard), "
            "S20→Excel factory (openpyxl), S21→reportes trazables (cierre CP-N2-B). "
            "Full packages 8 subtopics each (theory+demo+E1/E2/E3, 2 hints) = 8 demos + 24 exercises each. "
            "Platform ids data-engineering / databases-orm / rag / fastapi preserved. "
            "All demos/solutions executed with python3; UNVERIFIED=[]. Español peruano; synthetic data only. "
            "No seed/checkpoint/ledger edits."
        ),
        "blockers": [],
        "do_not_edit": [
            "course-state/checkpoint.json",
            "course-state/section_ledger.json",
            "course-state/issue_registry.json",
            "course-state/parallel_orchestration.json",
            "prisma/seed.ts",
        ],
        "next_action": (
            "PHASE 5 exam banks for data-engineering, databases-orm, rag, fastapi V3 slugs. "
            "Do not mark S18–S21 passed from this lane."
        ),
        "verified_counts": {
            "S18": len(results["S18"]["log"]),
            "S19": len(results["S19"]["log"]),
            "S20": len(results["S20"]["log"]),
            "S21": len(results["S21"]["log"]),
            "UNVERIFIED": [],
        },
    }
    for s in ("S18", "S19", "S20", "S21"):
        lane[s].pop("fname", None)

    LANES.mkdir(parents=True, exist_ok=True)
    (LANES / f"{LANE_ID}.status.json").write_text(
        json.dumps(lane, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print("Wrote progress + lane status")
    print(
        "Verified counts:",
        {k: lane["verified_counts"][k] for k in ("S18", "S19", "S20", "S21", "UNVERIFIED")},
    )


if __name__ == "__main__":
    main()
